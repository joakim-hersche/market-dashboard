import io
from datetime import datetime
from typing import cast

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

# ── Design tokens — JP Morgan / institutional palette ─────────────────────────

_C_NAVY       = "003087"   # primary navy (JP Morgan)
_C_NAVY_DARK  = "001E5C"   # cover banner
_C_NAVY_SUB   = "002579"   # cover sub-banner
_C_GOLD       = "B8962E"   # gold accent
_C_WHITE      = "FFFFFF"
_C_ALT        = "F3F6FA"   # alternating row tint
_C_TOTAL_BG   = "E8EDF5"   # total/subtotal row
_C_INPUT_BG   = "EBF3FB"   # editable input fill
_C_INPUT_FG   = "1A4FA0"   # editable input font
_C_GRID       = "BFC9D4"   # cell border
_C_NOTE       = "5A6472"   # secondary/note text
_C_POS_BG     = "E8F5E9"   # positive fill
_C_NEG_BG     = "FBE9E9"   # negative fill
_C_AMBER_BG   = "FFF8E1"

# ── Chart palette — matches charts.py FALLBACK_COLORS (hex, no leading #) ─────
_CHART_COLORS = [
    "1D4ED8", "0EA5E9", "6366F1", "10B981", "F59E0B",
    "EC4899", "8B5CF6", "06B6D4", "22C55E", "F97316",
]

# ── Reusable style objects ─────────────────────────────────────────────────────

_HEADER_FILL   = PatternFill("solid", fgColor=_C_NAVY)
_HEADER_FONT   = Font(bold=True, color=_C_WHITE, size=10)
_LABEL_FONT    = Font(bold=True, size=10)
_TOTAL_FONT    = Font(bold=True, color=_C_NAVY, size=10)
_NOTE_FONT     = Font(italic=True, size=9, color=_C_NOTE)
_INPUT_FONT    = Font(color=_C_INPUT_FG, size=10)
_INPUT_FILL    = PatternFill("solid", fgColor=_C_INPUT_BG)
_ALT_FILL      = PatternFill("solid", fgColor=_C_ALT)
_TOTAL_FILL    = PatternFill("solid", fgColor=_C_TOTAL_BG)
_TEMPLATE_FILL = PatternFill("solid", fgColor=_C_INPUT_BG)
_GREEN_FILL    = PatternFill("solid", fgColor=_C_POS_BG)
_RED_FILL      = PatternFill("solid", fgColor=_C_NEG_BG)
_AMBER_FILL    = PatternFill("solid", fgColor=_C_AMBER_BG)

_CELL_BORDER = Border(bottom=Side(style="thin", color=_C_GRID))
_TOP_BORDER  = Border(top=Side(style="medium", color=_C_NAVY))

_CURRENCY_FMT = {
    "USD": '"$"#,##0.00',
    "EUR": '"€"#,##0.00',
    "GBP": '"£"#,##0.00',
    "CHF": '"CHF "#,##0.00',
}

_TABLE_STYLE = TableStyleInfo(
    name="TableStyleMedium9",
    showFirstColumn=False, showLastColumn=False,
    showRowStripes=False,  showColumnStripes=False,
)

# ── Helpers ────────────────────────────────────────────────────────────────────

def _write_headers(ws, headers: list[str], start_row: int = 1) -> None:
    for col, h in enumerate(headers, 1):
        cell           = ws.cell(start_row, col, h)
        cell.fill      = _HEADER_FILL
        cell.font      = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = Border(bottom=Side(style="medium", color=_C_GOLD))
    ws.row_dimensions[start_row].height = 22


def _autofit(ws) -> None:
    for col in ws.columns:
        width = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(width + 3, 42)


def _add_table(ws, name: str, ref: str) -> None:
    t = Table(displayName=name, ref=ref)
    t.tableStyleInfo = _TABLE_STYLE
    ws.add_table(t)


def _set_print(ws, area: str | None = None) -> None:
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    if area:
        ws.print_area = area


def _no_gridlines(ws) -> None:
    ws.sheet_view.showGridLines = False


def _row_fill(row_idx: int) -> PatternFill | None:
    """Return alternating row fill for data rows (even = light tint, odd = none)."""
    return _ALT_FILL if row_idx % 2 == 0 else None


# ── Sheet builders ─────────────────────────────────────────────────────────────

def _sheet_net_worth(wb: Workbook, kpis: dict, currency: str) -> None:
    ws       = wb.create_sheet("Net Worth")
    curr_fmt = _CURRENCY_FMT.get(currency, "#,##0.00")
    n_cols   = 9

    _no_gridlines(ws)

    # ── Row 1: primary header banner ──────────────────────────────────────────
    ws.merge_cells(f"A1:{get_column_letter(n_cols)}1")
    h           = ws["A1"]
    h.value     = "PORTFOLIO REPORT"
    h.font      = Font(bold=True, size=20, color=_C_WHITE)
    h.fill      = PatternFill("solid", fgColor=_C_NAVY_DARK)
    h.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[1].height = 52

    # ── Row 2: sub-header ─────────────────────────────────────────────────────
    ws.merge_cells(f"A2:{get_column_letter(n_cols)}2")
    sub           = ws["A2"]
    sub.value     = (
        f"Prepared  {datetime.now().strftime('%d %B %Y')}    ·    "
        f"Reporting Currency: {currency}    ·    Private & Confidential"
    )
    sub.font      = Font(size=10, color="BFCDE0")
    sub.fill      = PatternFill("solid", fgColor=_C_NAVY_SUB)
    sub.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[2].height = 26

    # ── Row 3: gold accent line ───────────────────────────────────────────────
    ws.merge_cells(f"A3:{get_column_letter(n_cols)}3")
    ws["A3"].fill = PatternFill("solid", fgColor=_C_GOLD)
    ws.row_dimensions[3].height = 5

    # ── Row 4: spacer ─────────────────────────────────────────────────────────
    ws.row_dimensions[4].height = 16

    # ── Rows 5–7: KPI cards ───────────────────────────────────────────────────
    kpi_cards = [
        ("A", "C", "TOTAL PORTFOLIO VALUE", "=Summary!B5",  curr_fmt),
        ("D", "F", "TODAY'S CHANGE",         "=Summary!B6",  curr_fmt),
        ("G", "I", "TOTAL RETURN (%)",       "=Summary!B10", '0.00"%"'),
    ]
    ws.row_dimensions[5].height = 18
    ws.row_dimensions[6].height = 38
    ws.row_dimensions[7].height = 10

    for start_col, end_col, label, value, fmt in kpi_cards:
        ws.merge_cells(f"{start_col}5:{end_col}5")
        lbl           = ws[f"{start_col}5"]
        lbl.value     = label
        lbl.font      = Font(size=9, color=_C_NOTE)
        lbl.alignment = Alignment(horizontal="left", vertical="bottom", indent=1)

        ws.merge_cells(f"{start_col}6:{end_col}6")
        val               = ws[f"{start_col}6"]
        val.value         = value
        val.number_format = fmt
        val.font          = Font(bold=True, size=18, color=_C_NAVY_DARK)
        val.alignment     = Alignment(horizontal="left", vertical="center", indent=1)

        ws.merge_cells(f"{start_col}7:{end_col}7")
        ws[f"{start_col}7"].border = Border(bottom=Side(style="thin", color=_C_GOLD))

    # ── Row 8: spacer ─────────────────────────────────────────────────────────
    ws.row_dimensions[8].height = 18

    # ── Row 9: section label ──────────────────────────────────────────────────
    ws.merge_cells(f"A9:{get_column_letter(n_cols)}9")
    sec           = ws["A9"]
    sec.value     = "NET WORTH BREAKDOWN"
    sec.font      = Font(bold=True, size=9, color=_C_NAVY_DARK)
    sec.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    sec.border    = Border(bottom=Side(style="thin", color=_C_NAVY))
    ws.row_dimensions[9].height = 22

    # ── Row 10: table header ──────────────────────────────────────────────────
    for col, label in enumerate(["Category", f"Value ({currency})", "% of Total"], 1):
        cell           = ws.cell(10, col, label)
        cell.fill      = _HEADER_FILL
        cell.font      = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = Border(bottom=Side(style="medium", color=_C_GOLD))
    ws.row_dimensions[10].height = 22

    # ── Rows 11–12: data ──────────────────────────────────────────────────────
    ws.cell(11, 1, "Portfolio (Dashboard)").font = Font(size=10)
    ws.cell(11, 2, "=Summary!B5").number_format  = curr_fmt
    ws.cell(11, 3, '=IF(B13=0,"",B11/B13*100)').number_format = '0.0"%"'

    ws.cell(12, 1, "Other Assets (Manual)").font = Font(size=10)
    ws.cell(12, 2, f"='Other Assets'!H{_OA_SUM_ROW}").number_format = curr_fmt
    ws.cell(12, 3, '=IF(B13=0,"",B12/B13*100)').number_format = '0.0"%"'

    for r in (11, 12):
        for col in range(1, 4):
            ws.cell(r, col).border = _CELL_BORDER
        ws.row_dimensions[r].height = 20

    # alternate fill on row 12
    for col in range(1, 4):
        ws.cell(12, col).fill = _ALT_FILL

    # ── Row 13: total ─────────────────────────────────────────────────────────
    for col in range(1, 4):
        c        = ws.cell(13, col)
        c.border = _TOP_BORDER
        c.fill   = _TOTAL_FILL
        c.font   = _TOTAL_FONT
    ws.cell(13, 1, "Total Net Worth")
    ws.cell(13, 2, "=B11+B12").number_format = curr_fmt
    ws.cell(13, 3, "100%").number_format      = '0.0"%"'
    ws.row_dimensions[13].height = 22

    # ── Row 14: spacer ────────────────────────────────────────────────────────
    ws.row_dimensions[14].height = 14

    # ── Bar chart ─────────────────────────────────────────────────────────────
    chart              = BarChart()
    chart.type         = "bar"
    chart.title        = "Net Worth Breakdown"
    chart.style        = 10
    chart.grouping     = "clustered"
    chart.x_axis.title = f"Value ({currency})"
    chart.y_axis.delete = True

    data = Reference(ws, min_col=2, max_col=2, min_row=10, max_row=12)
    cats = Reference(ws, min_col=1, min_row=11, max_row=12)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.series[0].graphicalProperties.solidFill = _CHART_COLORS[0]
    chart.series[0].graphicalProperties.line.solidFill = _CHART_COLORS[0]
    chart.width  = 18
    chart.height = 10
    ws.add_chart(chart, "E9")

    # ── Row 16: disclaimer ────────────────────────────────────────────────────
    ws.merge_cells(f"A16:{get_column_letter(n_cols)}16")
    disc           = ws["A16"]
    disc.value     = (
        "This report is generated from live market data and is for informational purposes only. "
        "Past performance is not indicative of future results. Not financial advice."
    )
    disc.font      = _NOTE_FONT
    disc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    ws.row_dimensions[16].height = 28

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 14
    for col in "DEFGHI":
        ws.column_dimensions[col].width = 16

    _set_print(ws, "A1:I16")


def _sheet_summary(wb: Workbook, kpis: dict, currency: str, n_rows: int) -> None:
    ws       = wb.create_sheet("Summary")
    curr_fmt = _CURRENCY_FMT.get(currency, "#,##0.00")

    _no_gridlines(ws)

    # ── Row 1: title ──────────────────────────────────────────────────────────
    ws.merge_cells("A1:B1")
    h           = ws["A1"]
    h.value     = "PORTFOLIO SUMMARY"
    h.font      = Font(bold=True, size=14, color=_C_NAVY_DARK)
    h.fill      = PatternFill("solid", fgColor=_C_ALT)
    h.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    h.border    = Border(bottom=Side(style="medium", color=_C_GOLD))
    ws.row_dimensions[1].height = 38

    # ── Rows 2–3: metadata ────────────────────────────────────────────────────
    ws.cell(2, 1, "Report Date").font        = _NOTE_FONT
    ws.cell(2, 1).alignment                  = Alignment(indent=1)
    ws.cell(2, 2, datetime.now().strftime("%d %B %Y  %H:%M"))
    ws.cell(2, 2).font                       = Font(size=10, color=_C_NAVY)
    ws.cell(2, 2).alignment                  = Alignment(horizontal="right")
    ws.cell(2, 2).border                     = _CELL_BORDER

    ws.cell(3, 1, "Reporting Currency").font = _NOTE_FONT
    ws.cell(3, 1).alignment                  = Alignment(indent=1)
    ws.cell(3, 2, currency)
    ws.cell(3, 2).font                       = Font(size=10, color=_C_NAVY)
    ws.cell(3, 2).alignment                  = Alignment(horizontal="right")
    ws.cell(3, 2).border                     = _CELL_BORDER
    for r in (2, 3):
        ws.row_dimensions[r].height = 20

    # ── Row 4: section separator ──────────────────────────────────────────────
    ws.merge_cells("A4:B4")
    sec           = ws["A4"]
    sec.value     = "PERFORMANCE"
    sec.font      = Font(bold=True, size=8, color=_C_NOTE)
    sec.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    sec.border    = Border(bottom=Side(style="thin", color=_C_GRID))
    ws.row_dimensions[4].height = 20

    # ── Rows 5–11: KPIs ───────────────────────────────────────────────────────
    # NOTE: row numbers here must stay in sync with Net Worth references to Summary!B5
    _A = get_column_letter(_POS_IDX["Ticker"])
    _CB = get_column_letter(_POS_IDX["Cost Basis"])
    _H = get_column_letter(_POS_IDX["Total Value"])
    _I = get_column_letter(_POS_IDX["Dividends"])
    _J = get_column_letter(_POS_IDX["Daily P&L"])
    formula_rows = [
        ("Total Portfolio Value", f"=SUM(Positions!${_H}$2:${_H}${n_rows + 1})",                        curr_fmt,    True),
        ("Today's Change",        f"=SUM(Positions!${_J}$2:${_J}${n_rows + 1})",                        curr_fmt,    True),
        ("Cost Basis",             f"=SUM(Positions!${_CB}$2:${_CB}${n_rows + 1})",                      curr_fmt,    True),
        ("Dividends Received",     f"=SUM(Positions!${_I}$2:${_I}${n_rows + 1})",                      curr_fmt,    True),
        ("Total Return",           "=B5+B8-B7",                                                  curr_fmt,    True),
        ("Total Return (%)",       '=IF(B7=0,"",B9/B7*100)',                                     '0.00"%"',  True),
        ("Unique Holdings",         f"=SUMPRODUCT(1/COUNTIF(Positions!${_A}$2:${_A}${n_rows + 1},Positions!${_A}$2:${_A}${n_rows + 1}))", "0", True),
    ]

    for idx, (label, value, fmt, is_formula) in enumerate(formula_rows, 5):
        lbl_cell           = ws.cell(idx, 1, label)
        lbl_cell.alignment = Alignment(vertical="center", indent=1)
        lbl_cell.border    = _CELL_BORDER
        if idx == 5:  # Total Portfolio Value — highlight
            lbl_cell.font = Font(bold=True, size=10, color=_C_NAVY)
            for c in (1, 2):
                ws.cell(idx, c).fill = _TOTAL_FILL
        else:
            lbl_cell.font = Font(size=10)

        val_cell               = ws.cell(idx, 2, value)
        val_cell.number_format = fmt
        val_cell.alignment     = Alignment(horizontal="right", vertical="center")
        val_cell.border        = _CELL_BORDER
        val_cell.font          = Font(size=10) if is_formula else _INPUT_FONT

        if idx % 2 == 0:
            ws.cell(idx, 1).fill = _ALT_FILL
            if not ws.cell(idx, 2).fill.fgColor.rgb == _C_TOTAL_BG:
                ws.cell(idx, 2).fill = _ALT_FILL

        ws.row_dimensions[idx].height = 22

    # Bold Total Portfolio Value value
    ws.cell(5, 2).font = Font(bold=True, size=10, color=_C_NAVY)

    # ── Risk section ─────────────────────────────────────────────────────────
    risk_start = 13
    ws.merge_cells(f"A{risk_start}:B{risk_start}")
    sec2           = ws.cell(risk_start, 1, "RISK")
    sec2.font      = Font(bold=True, size=8, color=_C_NOTE)
    sec2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    sec2.border    = Border(bottom=Side(style="thin", color=_C_GRID))
    ws.row_dimensions[risk_start].height = 20

    # These are pre-computed values passed in kpis dict
    risk_rows = [
        ("Sharpe Ratio (portfolio)", kpis.get("portfolio_sharpe"),  "0.00"),
        ("Max Drawdown (portfolio)", kpis.get("portfolio_max_dd"),  '0.0"%"'),
        ("Annualized Volatility",    kpis.get("portfolio_vol"),     '0.0"%"'),
    ]
    _summary_risk_comment = (
        f"Computed by script on {datetime.now().strftime('%d %B %Y')}. "
        "Source: historical price data (yfinance). Re-run script to refresh."
    )
    for idx, (label, value, fmt) in enumerate(risk_rows, risk_start + 1):
        lbl_cell           = ws.cell(idx, 1, label)
        lbl_cell.font      = Font(size=10)
        lbl_cell.alignment = Alignment(vertical="center", indent=1)
        lbl_cell.border    = _CELL_BORDER

        val_cell               = ws.cell(idx, 2, round(value, 2) if value is not None else None)
        val_cell.number_format = fmt
        val_cell.alignment     = Alignment(horizontal="right", vertical="center")
        val_cell.border        = _CELL_BORDER
        val_cell.font          = Font(size=10)
        if value is not None:
            val_cell.comment = Comment(_summary_risk_comment, "Portfolio Report")

        if idx % 2 == 0:
            ws.cell(idx, 1).fill = _ALT_FILL
            ws.cell(idx, 2).fill = _ALT_FILL

        ws.row_dimensions[idx].height = 22

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 22
    _set_print(ws, f"A1:B{risk_start + len(risk_rows)}")


# Column layout for Positions (must be stable — other sheets reference by column position)
_POS_COLS = [
    "Ticker", "Company", "Purchase", "Shares",
    "Buy Price", "Purchase Date", "Current Price",
    "Target Price", "Upside (%)",
    "Total Value", "Dividends", "Daily P&L",
    "Return (%)", "Weight (%)", "Cost Basis",
]
# Map name → 1-based column index
_POS_IDX = {c: i for i, c in enumerate(_POS_COLS, 1)}


def _sheet_positions(wb: Workbook, df: pd.DataFrame, name_map: dict, currency: str,
                     target_prices: dict[str, float | None] | None = None) -> None:
    ws       = wb.create_sheet("Positions")
    curr_fmt = _CURRENCY_FMT.get(currency, "#,##0.00")

    _no_gridlines(ws)

    export_df = df.copy()
    export_df.insert(1, "Company", cast(pd.Series, export_df["Ticker"]).map(lambda x: name_map.get(x, x)))

    # Add Target Price column from fundamentals data
    _tp = target_prices or {}
    export_df["Target Price"] = export_df["Ticker"].map(lambda t: _tp.get(t))
    # Upside % is computed as formula in Excel, but we need a placeholder column
    export_df["Upside (%)"] = None

    export_df = cast(pd.DataFrame, export_df[[c for c in _POS_COLS if c in export_df.columns]])
    export_df = export_df.rename(columns={"Purchase": "Lot #"})

    # Column letters
    D = get_column_letter(_POS_IDX["Shares"])
    E = get_column_letter(_POS_IDX["Buy Price"])
    G = get_column_letter(_POS_IDX["Current Price"])
    TP = get_column_letter(_POS_IDX["Target Price"])
    UP = get_column_letter(_POS_IDX["Upside (%)"])
    H = get_column_letter(_POS_IDX["Total Value"])
    I = get_column_letter(_POS_IDX["Dividends"])
    K = get_column_letter(_POS_IDX["Return (%)"])
    L = get_column_letter(_POS_IDX["Weight (%)"])
    CB = get_column_letter(_POS_IDX["Cost Basis"])

    n          = len(export_df)
    last_data  = n + 1
    totals_row = last_data + 2

    _write_headers(ws, list(export_df.columns))

    # Blue-input columns: Shares, Buy Price, Current Price (hardcoded; FX-converted to base currency)
    input_cols = {"Shares", "Buy Price", "Current Price"}

    for row_idx, row in enumerate(export_df.itertuples(index=False), 2):
        alt = _row_fill(row_idx)
        for col_idx, (col_name, value) in enumerate(zip(export_df.columns, row), 1):
            cell      = ws.cell(row_idx, col_idx)
            safe      = value if pd.notna(value) else None
            cell.border = _CELL_BORDER
            if alt and col_name not in input_cols:
                cell.fill = alt

            if col_name == "Purchase Date":
                # Convert text dates (e.g. "2023-05-12") to actual date values
                # so Excel can sort, filter, and compute holding period.
                if isinstance(safe, str) and safe:
                    try:
                        safe = datetime.strptime(safe, "%Y-%m-%d").date()
                    except ValueError:
                        pass
                cell.value         = safe
                cell.number_format = "YYYY-MM-DD"
            elif col_name == "Current Price":
                # Hardcoded FX-converted price from positions_df (base currency).
                # Price History stores native-currency prices so INDEX/MATCH would
                # give wrong values for EUR/CHF/GBX tickers.
                cell.value         = safe
                cell.number_format = curr_fmt
            elif col_name == "Target Price":
                cell.value         = safe
                cell.number_format = curr_fmt
            elif col_name == "Upside (%)":
                # Formula: (Target Price - Current Price) / Current Price * 100
                cell.value = (
                    f'=IF(OR({TP}{row_idx}="",{G}{row_idx}=0),"",'
                    f'({TP}{row_idx}-{G}{row_idx})/{G}{row_idx}*100)'
                )
                cell.number_format = '0.00"%"'
                cell.font          = Font(size=10)
            elif col_name == "Total Value":
                cell.value         = f"={D}{row_idx}*{G}{row_idx}"
                cell.number_format = curr_fmt
                cell.font          = Font(size=10)
            elif col_name == "Return (%)":
                cell.value = (
                    f'=IF({E}{row_idx}*{D}{row_idx}=0,"",'
                    f'(({G}{row_idx}*{D}{row_idx}+{I}{row_idx}-{E}{row_idx}*{D}{row_idx})'
                    f'/({E}{row_idx}*{D}{row_idx}))*100)'
                )
                cell.number_format = '0.00"%"'
                cell.font          = Font(size=10)
            elif col_name == "Weight (%)":
                cell.value = (
                    f'=IF(SUM({H}$2:{H}${last_data})=0,"",'
                    f'{H}{row_idx}/SUM({H}$2:{H}${last_data})*100)'
                )
                cell.number_format = '0.00"%"'
                cell.font          = Font(size=10)
            elif col_name == "Cost Basis":
                # Formula so Cost Basis stays live if Shares or Buy Price inputs change
                cell.value         = f"={D}{row_idx}*{E}{row_idx}"
                cell.number_format = curr_fmt
                cell.font          = Font(size=10)
            else:
                cell.value = safe
                cell.font  = Font(size=10)
                if col_name in {"Buy Price", "Dividends", "Daily P&L"}:
                    cell.number_format = curr_fmt
                elif col_name == "Shares":
                    cell.number_format = "#,##0"

            if col_name in input_cols:
                cell.font = _INPUT_FONT
                cell.fill = _INPUT_FILL

        ws.row_dimensions[row_idx].height = 18

    # Conditional formatting: green/red on Return (%), Daily P&L, and Upside (%)
    J = get_column_letter(_POS_IDX["Daily P&L"])
    for col_l in (K, J, UP):
        rng = f"{col_l}2:{col_l}{last_data}"
        ws.conditional_formatting.add(rng, CellIsRule(operator="greaterThan", formula=["0"], fill=_GREEN_FILL))
        ws.conditional_formatting.add(rng, CellIsRule(operator="lessThan",    formula=["0"], fill=_RED_FILL))

    # Totals row
    ws.cell(totals_row, 1, "TOTAL").font = _TOTAL_FONT
    for col_idx in range(1, len(export_df.columns) + 1):
        c        = ws.cell(totals_row, col_idx)
        c.fill   = _TOTAL_FILL
        c.border = _TOP_BORDER
        c.font   = _TOTAL_FONT
    ws.row_dimensions[totals_row].height = 22

    sum_cols = {H: curr_fmt, I: curr_fmt, J: curr_fmt}
    for col_l, fmt in sum_cols.items():
        cell               = ws.cell(totals_row, _POS_IDX[{H: "Total Value", I: "Dividends", J: "Daily P&L"}[col_l]])
        cell.value         = f"=SUM({col_l}2:{col_l}{last_data})"
        cell.number_format = fmt
        cell.font          = _TOTAL_FONT

    ret_cell               = ws.cell(totals_row, _POS_IDX["Return (%)"])
    # True portfolio return: (total value + dividends - cost basis) / cost basis * 100
    # Cost basis is pre-computed in the Cost Basis column (split-adjusted)
    ret_cell.value         = (
        f'=IF(SUM({CB}2:{CB}{last_data})=0,"",'
        f'(SUM({H}2:{H}{last_data})+SUM({I}2:{I}{last_data})'
        f'-SUM({CB}2:{CB}{last_data}))'
        f'/SUM({CB}2:{CB}{last_data})*100)'
    )
    ret_cell.number_format = '0.00"%"'
    ret_cell.font          = _TOTAL_FONT

    wt_cell               = ws.cell(totals_row, _POS_IDX["Weight (%)"])
    wt_cell.value         = f"=SUM({L}2:{L}{last_data})"
    wt_cell.number_format = '0.00"%"'
    wt_cell.font          = _TOTAL_FONT

    _autofit(ws)
    ws.freeze_panes = "A2"
    _add_table(ws, "tblPositions", f"A1:{get_column_letter(len(export_df.columns))}{last_data}")
    _set_print(ws, f"A1:{get_column_letter(len(export_df.columns))}{totals_row}")


def _sheet_allocation(wb: Workbook, df: pd.DataFrame, name_map: dict, currency: str) -> None:
    ws       = wb.create_sheet("Allocation")
    curr_fmt = _CURRENCY_FMT.get(currency, "#,##0.00")

    _no_gridlines(ws)

    tickers  = list(dict.fromkeys(df["Ticker"]))
    n        = len(tickers)
    last_row = n + 1

    _write_headers(ws, ["Ticker", "Company", "Total Value", "Weight (%)"])

    _TV = get_column_letter(_POS_IDX["Total Value"])
    for row_idx, ticker in enumerate(tickers, 2):
        alt = _row_fill(row_idx)
        ws.cell(row_idx, 1, ticker).border  = _CELL_BORDER
        ws.cell(row_idx, 2, name_map.get(ticker, ticker)).border = _CELL_BORDER

        val_cell               = ws.cell(row_idx, 3,
            f"=SUMIF(Positions!$A:$A,A{row_idx},Positions!${_TV}:${_TV})")
        val_cell.number_format = curr_fmt
        val_cell.border        = _CELL_BORDER

        wt_cell               = ws.cell(row_idx, 4,
            f'=IF(SUM(C$2:C${last_row})=0,"",C{row_idx}/SUM(C$2:C${last_row})*100)')
        wt_cell.number_format = '0.00"%"'
        wt_cell.border        = _CELL_BORDER

        if alt:
            for col in range(1, 5):
                ws.cell(row_idx, col).fill = alt

        for col in range(1, 5):
            ws.cell(row_idx, col).font = Font(size=10)

        ws.row_dimensions[row_idx].height = 18

    # Horizontal bar chart
    chart              = BarChart()
    chart.type         = "bar"
    chart.title        = "Portfolio Allocation"
    chart.style        = 10
    chart.y_axis.title = "Ticker"
    chart.x_axis.title = "Weight (%)"

    data = Reference(ws, min_col=4, max_col=4, min_row=1, max_row=last_row)
    cats = Reference(ws, min_col=1, min_row=2, max_row=last_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.series[0].graphicalProperties.solidFill = _CHART_COLORS[0]
    chart.series[0].graphicalProperties.line.solidFill = _CHART_COLORS[0]
    chart.width  = 20
    chart.height = max(10, n * 1.2)
    ws.add_chart(chart, "F2")

    _autofit(ws)
    ws.freeze_panes = "A2"
    _add_table(ws, "tblAlloc", f"A1:D{last_row}")


def _sheet_risk(wb: Workbook, analytics_df: pd.DataFrame, name_map: dict, positions_df: pd.DataFrame,
                base_currency: str = "USD") -> None:
    ws = wb.create_sheet("Risk Metrics")

    _no_gridlines(ws)

    if analytics_df.empty:
        ws["A1"] = "Risk data not available — open the Risk & Analytics section in the dashboard first."
        return

    export_df = analytics_df.copy()
    export_df.insert(1, "Company", cast(pd.Series, export_df["Ticker"]).map(lambda x: name_map.get(x, x)))
    export_df = export_df.rename(columns={"Volatility": "Volatility (%)", "Max Drawdown": "Max Drawdown (%)"})
    headers   = ["Ticker", "Company", "Volatility (%)", "Max Drawdown (%)", "Sharpe Ratio", "Beta"]
    export_df = export_df[[c for c in headers if c in export_df.columns]]

    _write_headers(ws, headers)

    sharpe_col = None
    for col_idx, col_name in enumerate(headers, 1):
        if col_name == "Sharpe Ratio":
            sharpe_col = get_column_letter(col_idx)

    _risk_comment_text = (
        f"Computed by script on {datetime.now().strftime('%d %B %Y')}. "
        "Source: historical price data (yfinance). Re-run script to refresh."
    )
    for row_idx, row in enumerate(export_df.itertuples(index=False), 2):
        alt = _row_fill(row_idx)
        for col_idx, (col_name, value) in enumerate(zip(export_df.columns, row), 1):
            cell        = ws.cell(row_idx, col_idx, value if pd.notna(value) else None)
            cell.border = _CELL_BORDER
            cell.font   = Font(size=10)
            if alt:
                cell.fill = alt
            if col_name in ("Volatility (%)", "Max Drawdown (%)") and isinstance(value, (int, float)):
                cell.number_format = '0.0"%"'
                cell.comment = Comment(_risk_comment_text, "Portfolio Report")
            elif col_name in ("Sharpe Ratio", "Beta") and isinstance(value, (int, float)):
                cell.number_format = "0.00"
                cell.comment = Comment(_risk_comment_text, "Portfolio Report")
        ws.row_dimensions[row_idx].height = 18

    # Portfolio weighted-average row — uses SUMPRODUCT formulas linked to Allocation weights
    if not positions_df.empty and len(export_df) > 1:
        last_row   = len(export_df) + 1
        totals_row = last_row + 2
        alloc_last = len(export_df) + 1  # Allocation data rows mirror Risk Metrics rows

        ws.cell(totals_row, 1, "Portfolio (wgt avg*)").font = _TOTAL_FONT
        for col_idx in range(1, len(headers) + 1):
            c        = ws.cell(totals_row, col_idx)
            c.fill   = _TOTAL_FILL
            c.border = _TOP_BORDER
            c.font   = _TOTAL_FONT
        ws.row_dimensions[totals_row].height = 22

        for col_name in ("Volatility (%)", "Max Drawdown (%)", "Sharpe Ratio", "Beta"):
            if col_name not in export_df.columns:
                continue
            col_idx  = headers.index(col_name) + 1
            col_l    = get_column_letter(col_idx)
            # VLOOKUP each Risk ticker (column A) into Allocation!A:D to get its
            # weight, then compute the weighted average only over tickers present
            # in analytics. This correctly handles cases where some tickers are
            # excluded from analytics (insufficient history), so weights don't
            # need to sum to 100% across the full portfolio.
            formula = (
                f"=IFERROR("
                f"SUMPRODUCT("
                f"IFERROR(VLOOKUP(A2:A{last_row},Allocation!$A:$D,4,0),0),"
                f"IF(ISNUMBER({col_l}2:{col_l}{last_row}),{col_l}2:{col_l}{last_row},0)"
                f")/SUM(IFERROR(VLOOKUP(A2:A{last_row},Allocation!$A:$D,4,0),0))"
                f',"")'
            )
            cell               = ws.cell(totals_row, col_idx, formula)
            cell.font          = _TOTAL_FONT
            cell.number_format = '0.0"%"' if col_name in ("Volatility (%)", "Max Drawdown (%)") else "0.00"

    if sharpe_col:
        n_rows     = len(export_df) + 1
        data_range = f"{sharpe_col}2:{sharpe_col}{n_rows}"
        ws.conditional_formatting.add(data_range, CellIsRule(operator="greaterThanOrEqual", formula=["1"],      fill=_GREEN_FILL))
        ws.conditional_formatting.add(data_range, CellIsRule(operator="between",            formula=["0", "1"], fill=_AMBER_FILL))
        ws.conditional_formatting.add(data_range, CellIsRule(operator="lessThan",           formula=["0"],      fill=_RED_FILL))

    # Methodology note — risk-free rate source + benchmark used
    from src.risk_free import risk_free_label
    _BENCH_LABEL = {
        "USD": "S&P 500 (SPY)",
        "CHF": "SMI (^SSMI)",
        "EUR": "Euro Stoxx 50 (^STOXX50E)",
        "GBP": "FTSE 100 (^FTSE)",
        "SEK": "OMX Stockholm 30 (^OMX)",
    }
    rf_label = risk_free_label(base_currency)
    bench_label = _BENCH_LABEL.get(base_currency, "S&P 500 (SPY)")

    note_row = (len(export_df) + 4) if not positions_df.empty and len(export_df) > 1 else len(export_df) + 2
    note_cell = ws.cell(note_row, 1,
        f"Sharpe Ratio uses {rf_label} as the risk-free rate. "
        f"Beta is measured against {bench_label}. "
        f"*Weighted averages are position-weighted means of individual metrics. "
        "Max Drawdown is not additive — the portfolio max drawdown (see Summary) may differ from this weighted average. "
        f"Risk data last computed: {datetime.now().strftime('%d %B %Y  %H:%M')}. "
        "Re-download to refresh.")
    note_cell.font      = _NOTE_FONT
    note_cell.alignment = Alignment(wrap_text=True, indent=1)
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=len(headers))
    ws.row_dimensions[note_row].height = 36

    _autofit(ws)
    ws.freeze_panes = "A2"
    _add_table(ws, "tblRisk", f"A1:{get_column_letter(len(headers))}{len(export_df) + 1}")


def _sheet_attribution(wb: Workbook, positions_df: pd.DataFrame, name_map: dict) -> None:
    ws = wb.create_sheet("Attribution")

    _no_gridlines(ws)

    if positions_df.empty:
        ws["A1"] = "No positions available."
        return

    # Aggregate to ticker level
    grouped: pd.DataFrame = cast(pd.DataFrame, positions_df.groupby("Ticker", sort=False).agg(
        total_value=("Total Value", "sum"),
        cost_basis_sum=pd.NamedAgg(column="Cost Basis", aggfunc="sum"),
        dividends=("Dividends", "sum"),
    ))
    total_portfolio = grouped["total_value"].sum()
    if total_portfolio == 0:
        ws["A1"] = "No portfolio value to attribute."
        return

    grouped["weight"] = grouped["total_value"] / total_portfolio * 100
    grouped["return_pct"] = (
        (grouped["total_value"] + grouped["dividends"] - grouped["cost_basis_sum"])
        / grouped["cost_basis_sum"] * 100
    )
    grouped["contribution"] = grouped["weight"] / 100 * grouped["return_pct"]
    grouped = grouped.sort_values("contribution", ascending=False)

    headers = ["Ticker", "Company", "Weight (%)", "Return (%)", "Contribution (%)"]
    _write_headers(ws, headers)

    for row_idx, (ticker, row) in enumerate(grouped.iterrows(), 2):
        alt = _row_fill(row_idx)
        values = [ticker, name_map.get(ticker, ""), row["weight"], row["return_pct"], row["contribution"]]
        for col_idx, (col_name, value) in enumerate(zip(headers, values), 1):
            if col_name == "Contribution (%)":
                # Formula: (Weight% / 100) * (Return% / 100) * 100 = Weight% * Return% / 100
                cell = ws.cell(row_idx, col_idx, f"=C{row_idx}/100*D{row_idx}/100*100")
            else:
                cell = ws.cell(row_idx, col_idx, round(value, 2) if isinstance(value, float) else value)
            cell.border = _CELL_BORDER
            cell.font   = Font(size=10)
            if alt:
                cell.fill = alt
            if col_name in ("Weight (%)", "Return (%)", "Contribution (%)") and isinstance(value, (int, float)):
                cell.number_format = '0.00"%"'
            if col_name == "Contribution (%)":
                cell.number_format = '0.00"%"'

    # Total row
    last_data  = len(grouped) + 1
    totals_row = last_data + 1
    ws.cell(totals_row, 1, "Total").font = _TOTAL_FONT
    for col_idx in range(1, len(headers) + 1):
        c        = ws.cell(totals_row, col_idx)
        c.fill   = _TOTAL_FILL
        c.border = _TOP_BORDER
        c.font   = _TOTAL_FONT
    # Sum contribution column
    contrib_col = get_column_letter(5)
    ws.cell(totals_row, 5, f"=SUM({contrib_col}2:{contrib_col}{last_data})").font = _TOTAL_FONT
    ws.cell(totals_row, 5).number_format = '0.00"%"'

    # Conditional formatting on contribution column
    data_range = f"{contrib_col}2:{contrib_col}{last_data}"
    ws.conditional_formatting.add(data_range, CellIsRule(operator="greaterThan", formula=["0"], fill=_GREEN_FILL))
    ws.conditional_formatting.add(data_range, CellIsRule(operator="lessThan",    formula=["0"], fill=_RED_FILL))

    _autofit(ws)
    ws.freeze_panes = "A2"
    _add_table(ws, "tblAttribution", f"A1:{get_column_letter(len(headers))}{last_data}")


def _sheet_fundamentals(wb: Workbook, fund_rows: list[dict], name_map: dict) -> None:
    ws = wb.create_sheet("Fundamentals")

    _no_gridlines(ws)

    if not fund_rows:
        ws["A1"] = "Fundamentals data not available — open the Risk & Analytics section in the dashboard first."
        return

    fund_df = pd.DataFrame(fund_rows)
    fund_df.insert(1, "Company", cast(pd.Series, fund_df["Ticker"]).map(lambda x: name_map.get(x, x)))
    headers = ["Ticker", "Company", "Sector", "P/E Ratio", "Div Yield (%)", "1-Year Low", "1-Year High", "1-Year Position"]
    fund_df = fund_df[[c for c in headers if c in fund_df.columns]]

    _write_headers(ws, list(fund_df.columns))
    for row_idx, row in enumerate(fund_df.itertuples(index=False), 2):
        alt = _row_fill(row_idx)
        for col_idx, (col_name, value) in enumerate(zip(fund_df.columns, row), 1):
            cell        = ws.cell(row_idx, col_idx, value if pd.notna(value) else None)
            cell.border = _CELL_BORDER
            cell.font   = Font(size=10)
            if alt:
                cell.fill = alt
            if col_name == "P/E Ratio" and isinstance(value, (int, float)):
                cell.number_format = "0.0"
            elif col_name in ("Div Yield (%)", "1-Year Position") and isinstance(value, (int, float)):
                cell.number_format = '0.00"%"'
            elif col_name in ("1-Year Low", "1-Year High") and isinstance(value, (int, float)):
                cell.number_format = "#,##0.00"
        ws.row_dimensions[row_idx].height = 18

    _autofit(ws)
    ws.freeze_panes = "A2"
    _add_table(ws, "tblFund", f"A1:{get_column_letter(len(fund_df.columns))}{len(fund_df) + 1}")


def _sheet_correlation(wb: Workbook, price_histories: dict, positions_df: pd.DataFrame) -> None:
    ws = wb.create_sheet("Correlation")

    _no_gridlines(ws)

    returns = {
        t: h["Close"].pct_change().dropna()
        for t, h in price_histories.items()
        if not h.empty and "Close" in h.columns
    }
    if len(returns) < 2:
        ws["A1"] = "At least 2 positions with price history are needed to compute correlation."
        return

    corr_df = pd.DataFrame(returns).dropna().corr()
    tickers = list(corr_df.columns)

    # weights are linked via formula to Allocation sheet (VLOOKUP) so they stay in sync

    # Row 1: note
    ws.merge_cells(f"A1:{get_column_letter(len(tickers) + 2)}1")
    ws["A1"]      = "Correlation of daily returns (6-month). Weight (%) = portfolio allocation per ticker."
    ws["A1"].font = _NOTE_FONT

    # Row 2: column headers
    for col_idx, (label, fill) in enumerate(
        [("Ticker", True), ("Weight (%)", True)] + [(t, True) for t in tickers], 1
    ):
        cell           = ws.cell(2, col_idx, label if col_idx <= 2 else tickers[col_idx - 3])
        cell.fill      = _HEADER_FILL
        cell.font      = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border    = Border(bottom=Side(style="medium", color=_C_GOLD))
    ws.row_dimensions[2].height = 22

    # Data rows start at row 3
    for row_idx, t_row in enumerate(tickers, 3):
        alt = _row_fill(row_idx)

        label           = ws.cell(row_idx, 1, t_row)
        label.fill      = _HEADER_FILL
        label.font      = _HEADER_FONT
        label.alignment = Alignment(horizontal="center")
        label.border    = _CELL_BORDER

        wt_cell               = ws.cell(row_idx, 2,
            f'=IFERROR(VLOOKUP(A{row_idx},Allocation!$A:$D,4,FALSE),0)')
        wt_cell.number_format = '0.0"%"'
        wt_cell.alignment     = Alignment(horizontal="center")
        wt_cell.border        = _CELL_BORDER
        wt_cell.font          = Font(size=10)
        if alt:
            wt_cell.fill = alt

        for col_idx, t_col in enumerate(tickers, 3):
            cell               = ws.cell(row_idx, col_idx, round(corr_df.loc[t_row, t_col], 4))
            cell.number_format = "0.00"
            cell.alignment     = Alignment(horizontal="center")
            cell.border        = _CELL_BORDER
            cell.font          = Font(size=10)
            if alt:
                cell.fill = alt

        ws.row_dimensions[row_idx].height = 18

    # Color scale on correlation values
    n          = len(tickers)
    data_range = f"C3:{get_column_letter(n + 2)}{n + 2}"
    ws.conditional_formatting.add(data_range, ColorScaleRule(
        start_type="num", start_value=-1, start_color="C0392B",
        mid_type="num",   mid_value=0,    mid_color="FFFFFF",
        end_type="num",   end_value=1,    end_color="2E7D32",
    ))

    _autofit(ws)
    ws.freeze_panes = "C3"


def _sheet_price_history(wb: Workbook, price_histories: dict) -> None:
    ws = wb.create_sheet("Price History")

    _no_gridlines(ws)

    close_series = {
        t: (h["Close"] / 100 if t.endswith(".L") else h["Close"])
        for t, h in price_histories.items()
        if not h.empty and "Close" in h.columns
    }
    if not close_series:
        ws["A1"] = "No price history available."
        return

    pivot       = pd.DataFrame(close_series).sort_index()
    pivot.index = pd.to_datetime(pivot.index).tz_localize(None)
    pivot       = pivot.ffill()
    n_tickers   = len(pivot.columns)
    n_rows      = len(pivot)

    # Row 1: note about native currency
    ws.merge_cells(f"A1:{get_column_letter(n_tickers + 1)}1")
    note_ph           = ws["A1"]
    note_ph.value     = (
        "Prices are in each ticker's native trading currency (USD, EUR, GBP/100 for London, CHF). "
        "They are NOT converted to the reporting currency. "
        "This differs from the Positions sheet which shows FX-converted values."
    )
    note_ph.font      = _NOTE_FONT
    note_ph.fill      = PatternFill("solid", fgColor=_C_AMBER_BG)
    note_ph.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=2)
    ws.row_dimensions[1].height = 36

    ws.cell(2, 1, "Date").fill      = _HEADER_FILL
    ws.cell(2, 1).font              = _HEADER_FONT
    ws.cell(2, 1).alignment         = Alignment(horizontal="center")
    ws.cell(2, 1).border            = Border(bottom=Side(style="medium", color=_C_GOLD))
    for col_idx, t in enumerate(pivot.columns, 2):
        cell           = ws.cell(2, col_idx, t)
        cell.fill      = _HEADER_FILL
        cell.font      = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border    = Border(bottom=Side(style="medium", color=_C_GOLD))
    ws.row_dimensions[2].height = 22

    for row_idx, (date, row) in enumerate(pivot.iterrows(), 3):
        alt           = _row_fill(row_idx)
        date_cell     = ws.cell(row_idx, 1, cast(pd.Timestamp, date).to_pydatetime())
        date_cell.number_format = "YYYY-MM-DD"
        date_cell.font          = Font(size=10)
        date_cell.border        = _CELL_BORDER
        if alt:
            date_cell.fill = alt

        for col_idx, value in enumerate(row, 2):
            cell               = ws.cell(row_idx, col_idx, round(float(value), 4) if pd.notna(value) else None)
            cell.number_format = "#,##0.00"
            cell.font          = Font(size=10)
            cell.border        = _CELL_BORDER
            if alt:
                cell.fill = alt

        ws.row_dimensions[row_idx].height = 16

    # Line chart (header now at row 2, data at rows 3 to n_rows+2)
    chart              = LineChart()
    chart.title        = "Price History (6 months, native currency per ticker)"
    chart.style        = 10
    chart.y_axis.title = "Price (native currency)"
    chart.x_axis.title = "Date"
    chart.smooth       = True

    for col_idx in range(2, n_tickers + 2):
        data = Reference(ws, min_col=col_idx, max_col=col_idx, min_row=2, max_row=n_rows + 2)
        chart.add_data(data, titles_from_data=True)
        i = col_idx - 2
        ser = chart.series[i]
        ser.graphicalProperties.line.solidFill = _CHART_COLORS[i % len(_CHART_COLORS)]
        ser.graphicalProperties.line.width = 20000  # 2pt in EMU units
        ser.smooth = True

    cats = Reference(ws, min_col=1, min_row=3, max_row=n_rows + 2)
    chart.set_categories(cats)
    chart.width  = 28
    chart.height = 16
    ws.add_chart(chart, f"A{n_rows + 4}")

    _autofit(ws)
    ws.freeze_panes = "B3"


def _sheet_daily_returns(wb: Workbook, price_histories: dict) -> None:
    ws = wb.create_sheet("Daily Returns")

    _no_gridlines(ws)

    close_series = {
        t: h["Close"]
        for t, h in price_histories.items()
        if not h.empty and "Close" in h.columns
    }
    if not close_series:
        ws["A1"] = "No price history available."
        return

    pivot   = pd.DataFrame(close_series).sort_index()
    pivot.index = pd.to_datetime(pivot.index).tz_localize(None)
    # Do NOT ffill before pct_change: carrying a stale price forward produces a
    # spurious 0% return on non-trading days and understates daily volatility.
    # dropna(how="all") removes only rows where every ticker has no data.
    returns = pivot.pct_change(fill_method=None).dropna(how="all")

    ws.cell(1, 1, "Date").fill      = _HEADER_FILL
    ws.cell(1, 1).font              = _HEADER_FONT
    ws.cell(1, 1).alignment         = Alignment(horizontal="center")
    ws.cell(1, 1).border            = Border(bottom=Side(style="medium", color=_C_GOLD))
    for col_idx, t in enumerate(returns.columns, 2):
        cell           = ws.cell(1, col_idx, t)
        cell.fill      = _HEADER_FILL
        cell.font      = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border    = Border(bottom=Side(style="medium", color=_C_GOLD))
    ws.row_dimensions[1].height = 22

    for row_idx, (date, row) in enumerate(returns.iterrows(), 2):
        date_cell               = ws.cell(row_idx, 1, cast(pd.Timestamp, date).to_pydatetime())
        date_cell.number_format = "YYYY-MM-DD"
        date_cell.font          = Font(size=10)
        date_cell.border        = _CELL_BORDER
        for col_idx, value in enumerate(row, 2):
            safe               = round(float(value), 6) if pd.notna(value) else None
            cell               = ws.cell(row_idx, col_idx, safe)
            cell.number_format = '0.00%'  # standard % format: multiplies by 100 automatically
            cell.font          = Font(size=10)
            cell.border        = _CELL_BORDER
            if safe is not None:
                if safe > 0:
                    cell.fill = _GREEN_FILL
                elif safe < 0:
                    cell.fill = _RED_FILL

        ws.row_dimensions[row_idx].height = 16

    _autofit(ws)
    ws.freeze_panes = "B2"


# Other Assets layout constants
_OA_HEADERS = [
    "Asset Name", "Type", "Units / Shares",
    "Cost Basis", "Purchase Date", "Current Value",
    "Debt / Mortgage", "Net Equity", "Annual Income",
    "Unrealised Gain", "Weight (%)", "Notes",
]
_OA_DATA_ROWS   = range(3, 22)
_OA_SUM_ROW     = 22
_OA_COL_COST    = 4   # D
_OA_COL_VALUE   = 6   # F
_OA_COL_DEBT    = 7   # G
_OA_COL_EQUITY  = 8   # H
_OA_COL_INCOME  = 9   # I
_OA_COL_GAIN    = 10  # J
_OA_COL_WEIGHT  = 11  # K


def _sheet_other_assets(wb: Workbook, currency: str) -> None:
    ws       = wb.create_sheet("Other Assets")
    curr_fmt = _CURRENCY_FMT.get(currency, "#,##0.00")
    n_cols   = len(_OA_HEADERS)
    sum_row  = _OA_SUM_ROW

    _no_gridlines(ws)

    # Row 1: instruction banner
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    instruction           = ws.cell(1, 1,
        f"Add your non-market assets below. Values in {currency}. "
        "Blue cells are editable inputs. Orange cells are computed automatically."
    )
    instruction.fill      = PatternFill("solid", fgColor=_C_NAVY_DARK)
    instruction.font      = Font(italic=True, size=10, color="BFCDE0")
    instruction.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=2)
    ws.row_dimensions[1].height = 32

    # Row 2: column headers
    _write_headers(ws, _OA_HEADERS, start_row=2)
    ws.freeze_panes = "A3"

    # Rows 3–21: data rows
    for row_idx in _OA_DATA_ROWS:
        alt = _row_fill(row_idx)

        for col_idx in range(1, n_cols + 1):
            ws.cell(row_idx, col_idx).border = _CELL_BORDER
            ws.cell(row_idx, col_idx).font   = Font(size=10)
            if alt:
                ws.cell(row_idx, col_idx).fill = alt

        for col_idx in (_OA_COL_COST, _OA_COL_VALUE, _OA_COL_DEBT, _OA_COL_INCOME):
            c               = ws.cell(row_idx, col_idx)
            c.font          = _INPUT_FONT
            c.fill          = _INPUT_FILL
            c.number_format = curr_fmt

        ws.cell(row_idx, 5).number_format = "YYYY-MM-DD"

        # IF(F="",0,F) - IF(G="",0,G) treats empty inputs as 0 so subtraction never
        # sees "" on either side, avoiding #VALUE! when only one cell has data.
        eq               = ws.cell(row_idx, _OA_COL_EQUITY,
            f'=IF(AND(F{row_idx}="",G{row_idx}=""),"",IF(F{row_idx}="",0,F{row_idx})-IF(G{row_idx}="",0,G{row_idx}))')
        eq.number_format = curr_fmt
        eq.font          = Font(color="E65100", size=10)

        gain               = ws.cell(row_idx, _OA_COL_GAIN,
            f'=IF(AND(F{row_idx}="",D{row_idx}=""),"",IF(F{row_idx}="",0,F{row_idx})-IF(D{row_idx}="",0,D{row_idx}))')
        gain.number_format = curr_fmt
        gain.font          = Font(color="E65100", size=10)

        # IFERROR handles the case where H{row_idx} is "" (returns "" from equity formula)
        # which would cause #VALUE! in the division.
        wt               = ws.cell(row_idx, _OA_COL_WEIGHT,
            f'=IFERROR(IF(OR(H{row_idx}="",SUM(H$3:H${sum_row - 1})=0),"",H{row_idx}/SUM(H$3:H${sum_row - 1})*100),"")')
        wt.number_format = '0.0"%"'
        wt.font          = Font(color="E65100", size=10)

        ws.row_dimensions[row_idx].height = 18

    # Row 22: SUM totals
    for col_idx in range(1, n_cols + 1):
        c        = ws.cell(sum_row, col_idx)
        c.fill   = _TOTAL_FILL
        c.font   = _TOTAL_FONT
        c.border = _TOP_BORDER
    ws.cell(sum_row, 1, "Total")
    ws.row_dimensions[sum_row].height = 22

    for col_idx, col_letter in [
        (_OA_COL_COST,   "D"),
        (_OA_COL_VALUE,  "F"),
        (_OA_COL_DEBT,   "G"),
        (_OA_COL_EQUITY, "H"),
        (_OA_COL_INCOME, "I"),
        (_OA_COL_GAIN,   "J"),
    ]:
        c               = ws.cell(sum_row, col_idx, f"=SUM({col_letter}3:{col_letter}{sum_row - 1})")
        c.number_format = curr_fmt
        c.font          = _TOTAL_FONT

    # Type dropdown
    dv = DataValidation(
        type="list",
        formula1=(
            '"Real Estate,Private Equity,Cash & Savings,'
            'Bonds & Fixed Income,Crypto,Vehicles,'
            'Art & Collectibles,Business Ownership,'
            'Retirement Account,Other"'
        ),
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="Invalid Type",
        error="Select a type from the dropdown list.",
    )
    ws.add_data_validation(dv)
    dv.sqref = f"B3:B{sum_row - 1}"

    dv_num = DataValidation(
        type="decimal", operator="greaterThanOrEqual", formula1="0",
        showErrorMessage=True, errorStyle="warning",
        errorTitle="Invalid Value",
        error="Expected a number >= 0. You can still proceed.",
    )
    ws.add_data_validation(dv_num)
    dv_num.sqref = (
        f"D3:D{sum_row - 1} F3:F{sum_row - 1} "
        f"G3:G{sum_row - 1} I3:I{sum_row - 1}"
    )

    _add_table(ws, "tblOther", f"A2:{get_column_letter(n_cols)}{sum_row - 1}")

    _autofit(ws)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 22  # fits longest dropdown: "Retirement Account"
    ws.column_dimensions["L"].width = 25


def _sheet_currency_exposure(wb: Workbook, positions_df: pd.DataFrame, name_map: dict, currency: str) -> None:
    """Currency exposure breakdown showing portfolio allocation by trading currency."""
    from src.fx import get_ticker_currency

    ws = wb.create_sheet("Currency Exposure")
    curr_fmt = _CURRENCY_FMT.get(currency, "#,##0.00")
    _no_gridlines(ws)

    if positions_df.empty:
        ws["A1"] = "No positions available."
        return

    # Aggregate by trading currency
    ccy_data: dict[str, dict] = {}
    for ticker in positions_df["Ticker"].unique():
        ticker_df = positions_df[positions_df["Ticker"] == ticker]
        ticker_ccy = get_ticker_currency(ticker)
        if ticker_ccy == "GBX":
            ticker_ccy = "GBP"
        if ticker_ccy not in ccy_data:
            ccy_data[ticker_ccy] = {"value": 0.0, "tickers": []}
        ccy_data[ticker_ccy]["value"] += ticker_df["Total Value"].sum()
        ccy_data[ticker_ccy]["tickers"].append(ticker)

    total_value = sum(d["value"] for d in ccy_data.values())
    sorted_ccys = sorted(ccy_data.items(), key=lambda x: x[1]["value"], reverse=True)

    # Title
    ws.merge_cells("A1:E1")
    h           = ws["A1"]
    h.value     = "CURRENCY EXPOSURE"
    h.font      = Font(bold=True, size=14, color=_C_NAVY_DARK)
    h.fill      = PatternFill("solid", fgColor=_C_ALT)
    h.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    h.border    = Border(bottom=Side(style="medium", color=_C_GOLD))
    ws.row_dimensions[1].height = 38

    note = ws.cell(2, 1,
        f"Exposure by trading currency, converted to {currency} at spot rates. "
        "GBX positions are grouped under GBP.")
    note.font = _NOTE_FONT
    note.alignment = Alignment(indent=1)
    ws.merge_cells("A2:E2")
    ws.row_dimensions[2].height = 22

    headers = ["Currency", f"Value ({currency})", "Weight (%)", "Holdings", "Tickers"]
    _write_headers(ws, headers, start_row=3)

    for row_idx, (ccy, data) in enumerate(sorted_ccys, 4):
        alt = _row_fill(row_idx)
        weight = data["value"] / total_value * 100 if total_value > 0 else 0

        ws.cell(row_idx, 1, ccy).font = Font(bold=True, size=10)
        ws.cell(row_idx, 2, round(data["value"], 2)).number_format = curr_fmt
        ws.cell(row_idx, 2).font = Font(size=10)
        ws.cell(row_idx, 3, round(weight, 1)).number_format = '0.0"%"'
        ws.cell(row_idx, 3).font = Font(size=10)
        ws.cell(row_idx, 4, len(data["tickers"])).font = Font(size=10)
        ws.cell(row_idx, 5, ", ".join(data["tickers"])).font = Font(size=10)

        for c in range(1, 6):
            ws.cell(row_idx, c).border = _CELL_BORDER
            if alt:
                ws.cell(row_idx, c).fill = alt
        ws.row_dimensions[row_idx].height = 18

    last_data = 3 + len(sorted_ccys)
    _add_table(ws, "tblCurrencyExp", f"A3:E{last_data}")

    # Totals row — outside the table
    # Use =Summary!B5 so the total matches the portfolio value exactly (no FX rounding gap)
    totals_row = last_data + 1
    ws.cell(totals_row, 1, "TOTAL").font = _TOTAL_FONT
    ws.cell(totals_row, 2, "=Summary!B5").number_format = curr_fmt
    ws.cell(totals_row, 2).font = _TOTAL_FONT
    ws.cell(totals_row, 3, 100.0).number_format = '0.0"%"'
    ws.cell(totals_row, 3).font = _TOTAL_FONT
    ws.cell(totals_row, 4, len(positions_df["Ticker"].unique())).font = _TOTAL_FONT
    for c in range(1, 6):
        ws.cell(totals_row, c).fill = _TOTAL_FILL
        ws.cell(totals_row, c).border = _TOP_BORDER
    ws.row_dimensions[totals_row].height = 22

    # Pie chart
    chart              = BarChart()
    chart.type         = "bar"
    chart.title        = f"Currency Exposure ({currency})"
    chart.style        = 10
    chart.y_axis.delete = True
    chart.x_axis.title = "Weight (%)"

    data_ref = Reference(ws, min_col=3, max_col=3, min_row=3, max_row=last_data)
    cats = Reference(ws, min_col=1, min_row=4, max_row=last_data)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats)
    chart.series[0].graphicalProperties.solidFill = _CHART_COLORS[0]
    chart.series[0].graphicalProperties.line.solidFill = _CHART_COLORS[0]
    chart.width  = 18
    chart.height = max(8, len(sorted_ccys) * 1.5)
    ws.add_chart(chart, "G3")

    _autofit(ws)
    ws.column_dimensions["E"].width = 35


def _sheet_monte_carlo(
    wb: Workbook,
    bt_result: dict,
    ticker_mc_results: dict,
    portfolio_mc: dict,
    name_map: dict,
    currency: str,
) -> None:
    """
    Monte Carlo sheet with summary KPIs, two tables, and two line charts.

    Summary  — portfolio VaR(95%), CVaR(95%), diversification effect (1Y).
    Table 1  — Backtest: per-ticker hit rates, kurtosis, skewness, reliability.
    Table 2  — Forward projection: annualised return/volatility and p10/median/p90
               at 3M, 6M, and 1Y horizons per ticker, in base currency.
    Chart    — Simulated portfolio bands vs. actual over the past year.
    """
    ws = wb.create_sheet("Monte Carlo")
    _no_gridlines(ws)
    ccy_fmt = _CURRENCY_FMT.get(currency, '"$"#,##0.00')

    N_COLS = 13  # widest table (forward projection)

    if not bt_result and not ticker_mc_results and not portfolio_mc:
        ws["A1"] = (
            "Monte Carlo data not available. "
            "Each position needs at least 2 years of price history."
        )
        ws["A1"].font = _NOTE_FONT
        return

    row = 1

    # ── Title ─────────────────────────────────────────────────────────────────
    title = ws.cell(row, 1, "Monte Carlo Simulation")
    title.font = Font(bold=True, size=14, color=_C_NAVY)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=N_COLS)
    ws.row_dimensions[row].height = 26
    row += 1

    disc = ws.cell(
        row, 1,
        f"Statistical model based on historical log-return distributions calibrated on up to 5 years of data. "
        f"Assumes log-normally distributed daily returns and stable correlations — both simplifications. "
        f"Positions with excess kurtosis > 3 (fat-tailed) will have understated tail risk. "
        f"This is not financial advice.   Generated: {datetime.now().strftime('%d %B %Y  %H:%M')}.",
    )
    disc.font      = _NOTE_FONT
    disc.alignment = Alignment(wrap_text=True, indent=1)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=N_COLS)
    ws.row_dimensions[row].height = 40
    row += 2  # blank gap

    # ── Portfolio summary KPIs ────────────────────────────────────────────────
    if portfolio_mc and portfolio_mc.get("portfolio_paths") is not None:
        try:
            pp        = np.asarray(portfolio_mc["portfolio_paths"])   # (n_sims, horizon)
            pp_i      = np.asarray(portfolio_mc.get("portfolio_paths_i", pp))
            sv        = float(portfolio_mc.get("start_value", 0) or 0)
            end_corr  = pp[:, -1]
            end_indep = pp_i[:, -1]
            idx5      = max(1, int(0.05 * len(end_corr)))

            sorted_c  = np.sort(end_corr)
            var_abs   = sv - float(sorted_c[idx5])          # positive = loss
            cvar_abs  = sv - float(sorted_c[:idx5].mean())
            var_pct   = var_abs  / sv * 100 if sv else 0
            cvar_pct  = cvar_abs / sv * 100 if sv else 0

            corr_p10  = float(np.percentile(end_corr,  10))
            indep_p10 = float(np.percentile(end_indep, 10))
            div_abs   = indep_p10 - corr_p10     # +ve = correlation hurts

            sec0 = ws.cell(row, 1, "Portfolio Summary — 1-Year Forward Outlook")
            sec0.font = Font(bold=True, size=11, color=_C_NAVY)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=N_COLS)
            ws.row_dimensions[row].height = 20
            row += 1

            KPI_COLS = [
                ("Start Value",        sv,       ccy_fmt,   None),
                ("VaR 95%  (abs)",     var_abs,  ccy_fmt,   _RED_FILL),
                ("VaR 95%  (%)",       var_pct,  '0.0"%"',  _RED_FILL),
                ("CVaR 95%  (abs)",    cvar_abs, ccy_fmt,   _RED_FILL),
                ("CVaR 95%  (%)",      cvar_pct, '0.0"%"',  _RED_FILL),
                ("Diversification Effect  (p10 impact)", div_abs, ccy_fmt,
                 _AMBER_FILL if div_abs >= 0 else _GREEN_FILL),
            ]
            for col_idx, (label, val, fmt, fill) in enumerate(KPI_COLS, 1):
                lc = ws.cell(row, col_idx, label)
                lc.font = Font(bold=True, size=9, color=_C_NAVY)
                lc.fill = _HEADER_FILL if fill is None else PatternFill("solid", fgColor=_C_NAVY)
                lc.font = Font(bold=True, color=_C_WHITE, size=9)
                lc.alignment = Alignment(horizontal="center", wrap_text=True)
                ws.row_dimensions[row].height = 28
            row += 1
            for col_idx, (label, val, fmt, fill) in enumerate(KPI_COLS, 1):
                vc = ws.cell(row, col_idx, round(val, 2))
                vc.number_format = fmt
                vc.font          = Font(bold=True, size=11, color=_C_NAVY)
                vc.alignment     = Alignment(horizontal="center")
                vc.fill          = fill or _ALT_FILL
                vc.border        = _CELL_BORDER
            ws.row_dimensions[row].height = 22
            row += 1

            note_kpi = ws.cell(
                row, 1,
                f"Start Value = current portfolio in {currency}. "
                "VaR 95% = worst expected 1-year loss at 5th percentile of 1,000 simulations. "
                "CVaR 95% = average of the worst 5% outcomes (Expected Shortfall). "
                "Diversification Effect = difference between p10 under independent vs. correlated paths — "
                "positive means correlation amplifies downside risk.",
            )
            note_kpi.font      = _NOTE_FONT
            note_kpi.alignment = Alignment(wrap_text=True, indent=1)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=N_COLS)
            ws.row_dimensions[row].height = 36
            row += 2
        except Exception:
            row += 1  # silently skip if portfolio_mc is incomplete

    # ── Section 1: Backtest table ─────────────────────────────────────────────
    bt_section_start = row
    sec1 = ws.cell(row, 1, "Backtest — Model vs. Actual (Past 12 Months)")
    sec1.font = Font(bold=True, size=11, color=_C_NAVY)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws.row_dimensions[row].height = 20
    row += 1

    BT_HEADERS = [
        "Ticker", "Company",
        "Hit Rate 80% CI", "Hit Rate 50% CI",
        "Kurtosis", "Skewness", "Fat-tailed", "Reliability",
    ]
    _write_headers(ws, BT_HEADERS, start_row=row)
    ws.row_dimensions[row].height = 22
    bt_table_header_row = row
    row += 1

    bt_data_start = row
    if bt_result:
        for i, ticker in enumerate(bt_result.get("tickers_used", [])):
            hr   = bt_result["ticker_hit_rates"].get(ticker, {})
            flag = bt_result["ticker_flags"].get(ticker, {})
            hr_80    = hr.get("hit_rate_80")
            hr_50    = hr.get("hit_rate_50")
            kurtosis = flag.get("kurtosis")
            skewness = flag.get("skewness")
            fat      = "Yes" if flag.get("fat_tailed") else "No"
            if hr_80 is None:       rel = None
            elif hr_80 >= 80:       rel = "Good"
            elif hr_80 >= 65:       rel = "Moderate"
            else:                   rel = "Low"

            alt    = _row_fill(i + 1)
            values = [ticker, name_map.get(ticker, ticker), hr_80, hr_50, kurtosis, skewness, fat, rel]
            for col_idx, val in enumerate(values, 1):
                cell        = ws.cell(row, col_idx, val)
                cell.border = _CELL_BORDER
                cell.font   = Font(size=10)
                if alt:
                    cell.fill = alt
                if col_idx in (3, 4) and isinstance(val, (int, float)):
                    cell.number_format = '0.0"%"'
                elif col_idx in (5, 6) and isinstance(val, (int, float)):
                    cell.number_format = "0.00"
                # Colour coding overrides alternating fill
                if col_idx == 3 and isinstance(val, (int, float)):
                    cell.fill = _GREEN_FILL if val >= 80 else (_AMBER_FILL if val >= 65 else _RED_FILL)
                if col_idx == 5 and isinstance(val, (int, float)):
                    cell.fill = _GREEN_FILL if val <= 1  else (_AMBER_FILL if val <= 3  else _RED_FILL)
                if col_idx == 8 and isinstance(val, str):
                    cell.fill = {"Good": _GREEN_FILL, "Moderate": _AMBER_FILL, "Low": _RED_FILL}.get(val, PatternFill())

            ws.row_dimensions[row].height = 18
            row += 1

        bt_data_end = row - 1
        _add_table(ws, "tblMCBacktest",
                   f"A{bt_table_header_row}:{get_column_letter(len(BT_HEADERS))}{bt_data_end}")

        # Portfolio total row
        row += 1
        for col_idx in range(1, len(BT_HEADERS) + 1):
            c        = ws.cell(row, col_idx)
            c.fill   = _TOTAL_FILL
            c.border = _TOP_BORDER
            c.font   = _TOTAL_FONT
        ws.cell(row, 1, "Portfolio").font  = _TOTAL_FONT
        ws.cell(row, 1).fill              = _TOTAL_FILL
        ws.cell(row, 1).border            = _TOP_BORDER
        for col_idx, key in ((3, "hit_rate_80"), (4, "hit_rate_50")):
            v = bt_result.get(key)
            if v is not None:
                c                = ws.cell(row, col_idx, v)
                c.font           = _TOTAL_FONT
                c.number_format  = '0.0"%"'
                c.fill           = _GREEN_FILL if (key == "hit_rate_80" and v >= 80) or (key == "hit_rate_50") else _AMBER_FILL
                c.border         = _TOP_BORDER
        ws.row_dimensions[row].height = 22
        row += 1

        note_bt = ws.cell(row, 1,
            f"Trained on {bt_result.get('train_days', '?')} trading days before {bt_result.get('split_date', '?')}. "
            "Hit Rate = fraction of the past year's trading days where the actual portfolio value stayed within the "
            "simulated confidence band. Target: ~80% for the 80% band, ~50% for the 50% band.")
        note_bt.font      = _NOTE_FONT
        note_bt.alignment = Alignment(wrap_text=True, indent=1)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        ws.row_dimensions[row].height = 32
        row += 2
    else:
        no_bt = ws.cell(row, 1, "Backtest not available — each position needs at least 2 years of price history.")
        no_bt.font = _NOTE_FONT
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        row += 2

    # ── Section 2: Forward projection table ───────────────────────────────────
    sec2 = ws.cell(row, 1, "Forward Projection — Per Position")
    sec2.font = Font(bold=True, size=11, color=_C_NAVY)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=N_COLS)
    ws.row_dimensions[row].height = 20
    row += 1

    FWD_HEADERS = [
        "Ticker", "Company",
        "Ann. Return (%)", "Ann. Volatility (%)",
        "3M p10", "3M Median", "3M p90",
        "6M p10", "6M Median", "6M p90",
        "1Y p10",  "1Y Median",  "1Y p90",
    ]
    _write_headers(ws, FWD_HEADERS, start_row=row)
    ws.row_dimensions[row].height = 22
    fwd_table_header_row = row
    row += 1

    fwd_data_start = row
    if ticker_mc_results:
        for i, (ticker, mc) in enumerate(ticker_mc_results.items()):
            if not mc:
                continue
            pct = mc.get("percentiles")  # DataFrame indexed by future dates

            def _at(day_idx: int, col: str):
                if pct is None or len(pct) <= day_idx:
                    return None
                v = pct.iloc[day_idx][col]
                return float(v) if pd.notna(v) else None

            alt    = _row_fill(i + 1)
            values = [
                ticker, name_map.get(ticker, ticker),
                mc.get("mu_annual"), mc.get("sigma_annual"),
                _at(62, "p10"),  _at(62,  "p50"),  _at(62,  "p90"),
                _at(125, "p10"), _at(125, "p50"),  _at(125, "p90"),
                _at(251, "p10"), _at(251, "p50"),  _at(251, "p90"),
            ]
            for col_idx, val in enumerate(values, 1):
                cell        = ws.cell(row, col_idx, val)
                cell.border = _CELL_BORDER
                cell.font   = Font(size=10)
                if alt:
                    cell.fill = alt
                if col_idx in (3, 4) and isinstance(val, (int, float)):
                    cell.number_format = "0.0"
                elif col_idx >= 5 and isinstance(val, (int, float)):
                    cell.number_format = ccy_fmt
                if col_idx == 3 and isinstance(val, (int, float)):
                    cell.fill = _GREEN_FILL if val >= 0 else _RED_FILL
            ws.row_dimensions[row].height = 18
            row += 1

        fwd_data_end = row - 1
        _add_table(ws, "tblMCForward",
                   f"A{fwd_table_header_row}:{get_column_letter(len(FWD_HEADERS))}{fwd_data_end}")

    note_fwd = ws.cell(row, 1,
        f"Projected prices in {currency} using 1,000 simulations calibrated on up to 5 years of history. "
        "Ann. Return and Volatility are annualised from daily log-returns. "
        "p10/p90 = 80% confidence interval end-price. Assumes stable return distribution — "
        "verify against the Backtest hit rates above before drawing conclusions.")
    note_fwd.font      = _NOTE_FONT
    note_fwd.alignment = Alignment(wrap_text=True, indent=1)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=N_COLS)
    ws.row_dimensions[row].height = 36
    row += 3

    # ── Section 3: Backtest chart ─────────────────────────────────────────────
    if bt_result and "actual" in bt_result and "percentiles" in bt_result:
        actual = bt_result["actual"]
        pct_df = bt_result["percentiles"]

        # Write chart data starting at current row
        chart_data_start = row
        CHART_COLS = ["Date", "p10", "p25", "Median", "p75", "p90", "Actual"]
        for ci, h in enumerate(CHART_COLS, 1):
            cell      = ws.cell(chart_data_start, ci, h)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT

        for ri, (date, act_val) in enumerate(actual.items(), 1):
            r = chart_data_start + ri
            ws.cell(r, 1, date.to_pydatetime()).number_format = "YYYY-MM-DD"
            for ci_offset, col in enumerate(["p10", "p25", "p50", "p75", "p90"], 2):
                v = pct_df.loc[date, col] if date in pct_df.index else None
                ws.cell(r, ci_offset, round(float(v), 2) if v is not None and pd.notna(v) else None)
            ws.cell(r, 7, round(float(act_val), 2) if pd.notna(act_val) else None)

        n_chart_rows  = len(actual)
        chart_data_end = chart_data_start + n_chart_rows

        chart              = LineChart()
        chart.title        = "Backtest: Simulated Portfolio Bands vs. Actual"
        chart.style        = 10
        chart.y_axis.title = f"Portfolio Value ({currency})"
        chart.x_axis.title = "Date"
        chart.width        = 28
        chart.height       = 16

        SERIES_COLORS = [
            "C9D8F0",  # p10  — very light blue
            "7BAADE",  # p25  — light blue
            "2255A4",  # Median — medium blue
            "7BAADE",  # p75  — light blue (mirrors p25)
            "C9D8F0",  # p90  — very light blue (mirrors p10)
            "1A1A1A",  # Actual — near black
        ]
        SERIES_WIDTHS = [15000, 15000, 20000, 15000, 15000, 30000]  # EMU

        for col_idx in range(2, 8):  # columns B–G
            data = Reference(ws,
                min_col=col_idx, max_col=col_idx,
                min_row=chart_data_start, max_row=chart_data_end)
            chart.add_data(data, titles_from_data=True)
            ser = chart.series[col_idx - 2]
            ser.graphicalProperties.line.solidFill = SERIES_COLORS[col_idx - 2]
            ser.graphicalProperties.line.width     = SERIES_WIDTHS[col_idx - 2]
            ser.smooth = False

        # Median as dashed
        chart.series[2].graphicalProperties.line.dashDot = "dash"

        cats = Reference(ws, min_col=1, min_row=chart_data_start + 1, max_row=chart_data_end)
        chart.set_categories(cats)
        ws.add_chart(chart, f"A{chart_data_end + 2}")

    _autofit(ws)
    ws.freeze_panes = "A4"


def _sheet_scenario(wb: Workbook, positions_df: pd.DataFrame, name_map: dict, currency: str) -> None:
    ws       = wb.create_sheet("Scenario Analysis")
    curr_fmt = _CURRENCY_FMT.get(currency, "#,##0.00")

    _no_gridlines(ws)

    if positions_df.empty:
        ws["A1"] = "No positions available."
        return

    # Row 1: banner
    ws.merge_cells("A1:J1")
    banner           = ws["A1"]
    banner.value     = (
        "Scenario Analysis  —  edit Target Price (blue cells) to model portfolio impact. "
        "All other columns update automatically. "
        "Projected Price Return (%) measures price-only return from average cost basis to target price "
        "(excludes dividends; see Positions sheet Return % for total return including dividends)."
    )
    banner.font      = Font(italic=True, size=10, color="BFCDE0")
    banner.fill      = PatternFill("solid", fgColor=_C_NAVY_DARK)
    banner.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=2)
    ws.row_dimensions[1].height = 28

    headers = [
        "Ticker", "Company", "Total Shares", "Avg Buy Price",
        "Current Price", "Target Price",
        "Current Value", "Projected Value",
        "Value Change", "Projected Price Return (%)",
    ]
    _write_headers(ws, headers, start_row=2)
    ws.freeze_panes = "A3"

    tickers = list(dict.fromkeys(positions_df["Ticker"]))
    n       = len(tickers)

    agg = (
        positions_df.groupby("Ticker")
        .apply(lambda g: pd.Series({
            "Shares":        g["Shares"].sum(),
            "AvgBuyPrice":   g["Cost Basis"].sum() / g["Shares"].sum() if g["Shares"].sum() else 0,
            "CurrentPrice":  g["Current Price"].iloc[0],
            "TotalValue":    g["Total Value"].sum(),
        }), include_groups=False)
        .reindex(tickers)
    )

    last_data = n + 2  # header at row 2, data rows 3 to n+2

    _PD = get_column_letter(_POS_IDX["Shares"])
    _PE = get_column_letter(_POS_IDX["Buy Price"])
    _PTV = get_column_letter(_POS_IDX["Total Value"])

    for row_idx, ticker in enumerate(tickers, 3):
        row_data = agg.loc[ticker]
        alt      = _row_fill(row_idx)

        for col_idx in range(1, len(headers) + 1):
            c        = ws.cell(row_idx, col_idx)
            c.border = _CELL_BORDER
            c.font   = Font(size=10)
            if alt and col_idx not in (5, 6):  # don't overwrite input fills
                c.fill = alt

        ws.cell(row_idx, 1, ticker)
        ws.cell(row_idx, 2, name_map.get(ticker, ticker))

        ws.cell(row_idx, 3,
            f"=SUMIF(Positions!$A:$A,A{row_idx},Positions!${_PD}:${_PD})").number_format = "#,##0"

        ws.cell(row_idx, 4,
            f"=IFERROR(SUMPRODUCT((Positions!$A$2:$A$10000=A{row_idx})"
            f"*Positions!${_PD}$2:${_PD}$10000*Positions!${_PE}$2:${_PE}$10000)"
            f"/SUMIF(Positions!$A:$A,A{row_idx},Positions!${_PD}:${_PD}),\"\")"
        ).number_format = curr_fmt

        _G = get_column_letter(_POS_IDX["Current Price"])
        curr_cell               = ws.cell(row_idx, 5,
            f'=IFERROR(INDEX(Positions!${_G}:${_G},'
            f'MATCH(A{row_idx},Positions!$A:$A,0)),0)')
        curr_cell.number_format = curr_fmt
        curr_cell.font          = Font(size=10)
        curr_cell.border        = _CELL_BORDER

        tgt_cell               = ws.cell(row_idx, 6, float(row_data["CurrentPrice"]) if pd.notna(row_data["CurrentPrice"]) else 0)
        tgt_cell.number_format = curr_fmt
        tgt_cell.font          = _INPUT_FONT
        tgt_cell.fill          = _INPUT_FILL

        ws.cell(row_idx, 7,
            f"=SUMIF(Positions!$A:$A,A{row_idx},Positions!${_PTV}:${_PTV})"
        ).number_format = curr_fmt

        ws.cell(row_idx, 8, f"=C{row_idx}*F{row_idx}").number_format = curr_fmt

        change               = ws.cell(row_idx, 9, f"=H{row_idx}-G{row_idx}")
        change.number_format = curr_fmt

        ws.cell(row_idx, 10,
            f'=IF(D{row_idx}=0,"",((F{row_idx}-D{row_idx})/D{row_idx})*100)'
        ).number_format = '0.00"%"'

        ws.row_dimensions[row_idx].height = 18

    for col_l in ("I", "J"):
        rng = f"{col_l}3:{col_l}{last_data}"
        ws.conditional_formatting.add(rng, CellIsRule(operator="greaterThan", formula=["0"], fill=_GREEN_FILL))
        ws.conditional_formatting.add(rng, CellIsRule(operator="lessThan",    formula=["0"], fill=_RED_FILL))

    totals_row = last_data + 2
    ws.cell(totals_row, 1, "TOTAL").font = _TOTAL_FONT
    for col_idx in range(1, len(headers) + 1):
        c        = ws.cell(totals_row, col_idx)
        c.fill   = _TOTAL_FILL
        c.border = _TOP_BORDER
        c.font   = _TOTAL_FONT
    for col_l in ("G", "H", "I"):
        c               = ws.cell(totals_row, ord(col_l) - ord("A") + 1, f"=SUM({col_l}3:{col_l}{last_data})")
        c.number_format = curr_fmt
        c.font          = _TOTAL_FONT
    ws.row_dimensions[totals_row].height = 22

    _autofit(ws)
    ws.freeze_panes = "A3"
    _add_table(ws, "tblScenario", f"A2:{get_column_letter(len(headers))}{last_data}")


# ── Health Score sheet ────────────────────────────────────────────────────────

def _sheet_health(
    wb: Workbook,
    score_result: dict,
    findings: list[dict],
    sector_weights: dict[str, float],
    ticker_sector: dict[str, str],
) -> None:
    """Portfolio Health Score with pre-computed components, findings, and sector breakdown."""
    ws = wb.create_sheet("Health Score")
    _no_gridlines(ws)

    if not score_result or "total" not in score_result:
        ws["A1"] = "Health score data not available — open the Health tab in the dashboard first."
        return

    N_COLS = 6

    # ── Row 1: title ──────────────────────────────────────────────────────────
    row = 1
    ws.merge_cells(f"A{row}:{get_column_letter(N_COLS)}{row}")
    h           = ws["A1"]
    h.value     = "PORTFOLIO HEALTH SCORE"
    h.font      = Font(bold=True, size=14, color=_C_NAVY_DARK)
    h.fill      = PatternFill("solid", fgColor=_C_ALT)
    h.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    h.border    = Border(bottom=Side(style="medium", color=_C_GOLD))
    ws.row_dimensions[row].height = 38
    row += 1

    # Note: all health scores are script-computed — add a visible banner so readers
    # understand these are static values that require a re-export to refresh.
    ws.merge_cells(f"A{row}:{get_column_letter(N_COLS)}{row}")
    hs_note           = ws.cell(row, 1,
        f"All scores, findings, and sector weights were computed by the script on "
        f"{datetime.now().strftime('%d %B %Y  %H:%M')}. "
        "Re-download the report to refresh these values."
    )
    hs_note.font      = Font(italic=True, size=9, color="6B4C00")
    hs_note.fill      = PatternFill("solid", fgColor=_C_AMBER_BG)
    hs_note.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    ws.row_dimensions[row].height = 22
    row += 1

    # ── Row 2: overall score ─────────────────────────────────────────────────
    total_score = score_result["total"]
    if total_score >= 70:
        score_fill = _GREEN_FILL
    elif total_score >= 40:
        score_fill = _AMBER_FILL
    else:
        score_fill = _RED_FILL

    ws.merge_cells(f"A{row}:B{row}")
    lbl           = ws.cell(row, 1, "Overall Score")
    lbl.font      = Font(bold=True, size=11, color=_C_NAVY)
    lbl.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    val           = ws.cell(row, 3, round(total_score, 1))
    val.font      = Font(bold=True, size=18, color=_C_NAVY_DARK)
    val.number_format = "0.0"
    val.fill      = score_fill
    val.alignment = Alignment(horizontal="center", vertical="center")
    of_cell       = ws.cell(row, 4, "/ 100")
    of_cell.font  = Font(size=11, color=_C_NOTE)
    of_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 36
    row += 2

    # ── Score components table ───────────────────────────────────────────────
    ws.merge_cells(f"A{row}:{get_column_letter(N_COLS)}{row}")
    sec           = ws.cell(row, 1, "SCORE BREAKDOWN")
    sec.font      = Font(bold=True, size=9, color=_C_NAVY_DARK)
    sec.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    sec.border    = Border(bottom=Side(style="thin", color=_C_NAVY))
    ws.row_dimensions[row].height = 22
    row += 1

    comp_headers = ["Component", "Score", "Max", "Details"]
    _write_headers(ws, comp_headers, start_row=row)
    comp_header_row = row
    row += 1

    for comp in score_result["components"]:
        alt = _row_fill(row)
        ws.cell(row, 1, comp["name"]).font = Font(size=10)
        ws.cell(row, 2, round(comp["score"], 1)).number_format = "0.0"
        ws.cell(row, 2).font = Font(size=10)
        ws.cell(row, 3, comp["max_score"]).font = Font(size=10)
        ws.cell(row, 4, comp["details"]).font = Font(size=10)
        for c in range(1, 5):
            ws.cell(row, c).border = _CELL_BORDER
            if alt:
                ws.cell(row, c).fill = alt
        # Color the score cell
        pct = comp["score"] / comp["max_score"] if comp["max_score"] > 0 else 0
        if pct >= 0.7:
            ws.cell(row, 2).fill = _GREEN_FILL
        elif pct >= 0.4:
            ws.cell(row, 2).fill = _AMBER_FILL
        else:
            ws.cell(row, 2).fill = _RED_FILL
        ws.row_dimensions[row].height = 18
        row += 1

    _add_table(ws, "tblHealthComp", f"A{comp_header_row}:D{row - 1}")
    row += 1

    # ── Findings ─────────────────────────────────────────────────────────────
    if findings:
        ws.merge_cells(f"A{row}:{get_column_letter(N_COLS)}{row}")
        sec2           = ws.cell(row, 1, "FINDINGS")
        sec2.font      = Font(bold=True, size=9, color=_C_NAVY_DARK)
        sec2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        sec2.border    = Border(bottom=Side(style="thin", color=_C_NAVY))
        ws.row_dimensions[row].height = 22
        row += 1

        find_headers = ["Severity", "Finding", "Detail"]
        _write_headers(ws, find_headers, start_row=row)
        find_header_row = row
        row += 1

        _SEVERITY_FILL = {"red": _RED_FILL, "amber": _AMBER_FILL, "green": _GREEN_FILL}

        for f in findings:
            alt = _row_fill(row)
            sev_cell = ws.cell(row, 1, f["severity"].upper())
            sev_cell.font = Font(bold=True, size=10)
            sev_cell.fill = _SEVERITY_FILL.get(f["severity"], PatternFill())
            sev_cell.alignment = Alignment(horizontal="center")
            ws.cell(row, 2, f["headline"]).font = Font(bold=True, size=10)
            ws.cell(row, 3, f["body"]).font = Font(size=10)
            for c in range(1, 4):
                ws.cell(row, c).border = _CELL_BORDER
                if alt and c != 1:  # don't override severity fill
                    ws.cell(row, c).fill = alt
            ws.row_dimensions[row].height = 20
            row += 1

        _add_table(ws, "tblHealthFindings", f"A{find_header_row}:C{row - 1}")
        row += 1

    # ── Sector breakdown ─────────────────────────────────────────────────────
    if sector_weights:
        ws.merge_cells(f"A{row}:{get_column_letter(N_COLS)}{row}")
        sec3           = ws.cell(row, 1, "SECTOR EXPOSURE")
        sec3.font      = Font(bold=True, size=9, color=_C_NAVY_DARK)
        sec3.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        sec3.border    = Border(bottom=Side(style="thin", color=_C_NAVY))
        ws.row_dimensions[row].height = 22
        row += 1

        sect_headers = ["Sector", "Weight (%)", "Tickers"]
        _write_headers(ws, sect_headers, start_row=row)
        sect_header_row = row
        row += 1

        sorted_sectors = sorted(sector_weights.items(), key=lambda x: x[1], reverse=True)
        for sector, sw in sorted_sectors:
            alt = _row_fill(row)
            ws.cell(row, 1, sector).font = Font(size=10)
            ws.cell(row, 2, round(sw, 1)).number_format = '0.0"%"'
            ws.cell(row, 2).font = Font(size=10)
            # List tickers in this sector
            tickers_in = [t for t, s in ticker_sector.items() if s == sector]
            ws.cell(row, 3, ", ".join(tickers_in)).font = Font(size=10)
            for c in range(1, 4):
                ws.cell(row, c).border = _CELL_BORDER
                if alt:
                    ws.cell(row, c).fill = alt
            ws.row_dimensions[row].height = 18
            row += 1

        _add_table(ws, "tblHealthSectors", f"A{sect_header_row}:C{row - 1}")

        # Bar chart for sector weights
        n_sectors = len(sorted_sectors)
        chart              = BarChart()
        chart.type         = "bar"
        chart.title        = "Sector Exposure"
        chart.style        = 10
        chart.y_axis.delete = True
        chart.x_axis.title = "Weight (%)"

        data = Reference(ws, min_col=2, max_col=2, min_row=sect_header_row, max_row=sect_header_row + n_sectors)
        cats = Reference(ws, min_col=1, min_row=sect_header_row + 1, max_row=sect_header_row + n_sectors)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.series[0].graphicalProperties.solidFill = _CHART_COLORS[0]
        chart.series[0].graphicalProperties.line.solidFill = _CHART_COLORS[0]
        chart.width  = 20
        chart.height = max(8, n_sectors * 1.2)
        ws.add_chart(chart, f"E{sect_header_row}")

    _autofit(ws)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["D"].width = 30


# ── Income sheet ──────────────────────────────────────────────────────────────

def _sheet_income(wb: Workbook, positions_df: pd.DataFrame, fund_rows: list[dict],
                  name_map: dict, currency: str,
                  dividend_timeline: list[dict] | None = None,
                  portfolio: dict | None = None) -> None:
    """Per-position dividend income summary + forward dividend calendar."""
    from collections import Counter, defaultdict
    from src.fx import get_ticker_currency, get_fx_rate

    ws = wb.create_sheet("Income")
    curr_fmt = _CURRENCY_FMT.get(currency, '#,##0.00')
    headers = ["Ticker", "Company", "Shares", "Annual Income", "Yield (%)", "Yield on Cost (%)"]

    _no_gridlines(ws)

    # ── Section 1: title ─────────────────────────────────────────────────────
    ws.merge_cells("A1:F1")
    title           = ws["A1"]
    title.value     = "DIVIDEND INCOME"
    title.font      = Font(bold=True, size=14, color=_C_NAVY_DARK)
    title.fill      = PatternFill("solid", fgColor=_C_ALT)
    title.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    title.border    = Border(bottom=Side(style="medium", color=_C_GOLD))
    ws.row_dimensions[1].height = 38

    _write_headers(ws, headers, start_row=2)

    # Build a lookup from fund_rows
    fund_map = {}
    for fr in fund_rows:
        t = fr.get("Ticker")
        if t:
            fund_map[t] = fr

    row = 3
    tickers = sorted(positions_df["Ticker"].unique()) if not positions_df.empty else []
    for ticker in tickers:
        ticker_df = positions_df[positions_df["Ticker"] == ticker]
        total_shares = ticker_df["Shares"].sum()
        total_value = ticker_df["Total Value"].sum()
        cost_basis = ticker_df["Cost Basis"].sum()

        fund = fund_map.get(ticker, {})
        div_rate = fund.get("Dividend Rate")
        div_yield = fund.get("Div Yield (%)")

        div_ccy = fund.get("Financial Currency") or get_ticker_currency(ticker)
        fx_rate, _ = get_fx_rate(div_ccy, currency)

        # Compute per-share annual dividend in base currency so the Excel formula
        # =C{row} * per_share_annual_div remains auditable and stays live with Shares.
        per_share_annual_div = 0.0
        if div_rate and div_rate > 0:
            # Preferred: explicit per-share dividend rate from fundamentals data
            per_share_annual_div = div_rate * fx_rate
        elif div_yield and div_yield > 0 and total_shares > 0:
            # Fallback for ETFs (e.g. SPY, GLD) where Dividend Rate is not reported
            # but Div Yield (%) is known: derive per-share from current price × yield
            current_price_base = total_value / total_shares
            per_share_annual_div = current_price_base * div_yield / 100

        annual_income = per_share_annual_div * total_shares
        yoc = (annual_income / cost_basis * 100) if cost_basis > 0 else 0

        name = name_map.get(ticker, ticker)
        ws.cell(row, 1, ticker).font = Font(bold=True, size=10)
        ws.cell(row, 2, name)
        ws.cell(row, 3, total_shares).number_format = '#,##0'
        # Write as formula so the cell is auditable: =Shares * per_share_annual_div
        if per_share_annual_div > 0:
            ws.cell(row, 4, f"=C{row}*{per_share_annual_div:.6f}").number_format = curr_fmt
        else:
            ws.cell(row, 4, 0).number_format = curr_fmt
        ws.cell(row, 5, round(div_yield, 2) if div_yield else 0).number_format = '0.00"%"'
        ws.cell(row, 6, round(yoc, 2)).number_format = '0.00"%"'

        fill = _ALT_FILL if (row - 3) % 2 else None
        if fill:
            for c in range(1, len(headers) + 1):
                ws.cell(row, c).fill = fill
        for c in range(1, len(headers) + 1):
            ws.cell(row, c).border = _CELL_BORDER
        row += 1

    income_table_end = row - 1
    if income_table_end >= 3:
        _add_table(ws, "tblIncome", f"A2:{get_column_letter(len(headers))}{income_table_end}")

    # Totals row — sits OUTSIDE the table so other sheets' SUM/SUMIF won't capture it
    if tickers:
        totals_row = row
        ws.cell(totals_row, 1, "TOTAL").font = _TOTAL_FONT
        ws.cell(totals_row, 3, f"=SUM(C3:C{income_table_end})").number_format = '#,##0'
        ws.cell(totals_row, 3).font = _TOTAL_FONT
        ws.cell(totals_row, 4, f"=SUM(D3:D{income_table_end})").number_format = curr_fmt
        ws.cell(totals_row, 4).font = _TOTAL_FONT
        # Portfolio yield = total income / total portfolio value
        ws.cell(totals_row, 5,
            f'=IF(SUM(Positions!{get_column_letter(_POS_IDX["Total Value"])}2:'
            f'{get_column_letter(_POS_IDX["Total Value"])}{income_table_end + 100})=0,"",'
            f'D{totals_row}/SUM(Positions!{get_column_letter(_POS_IDX["Total Value"])}2:'
            f'{get_column_letter(_POS_IDX["Total Value"])}{income_table_end + 100})*100)'
        ).number_format = '0.00"%"'
        ws.cell(totals_row, 5).font = _TOTAL_FONT
        for c in range(1, len(headers) + 1):
            ws.cell(totals_row, c).fill = _TOTAL_FILL
            ws.cell(totals_row, c).border = _TOP_BORDER
        ws.row_dimensions[totals_row].height = 22
        row = totals_row + 1

    # ── Section 2: Forward dividend calendar ─────────────────────────────────
    if dividend_timeline and portfolio and tickers:
        row += 2

        ws.merge_cells(f"A{row}:N{row}")
        sec           = ws.cell(row, 1, "PROJECTED DIVIDEND CALENDAR (NEXT 12 MONTHS)")
        sec.font      = Font(bold=True, size=9, color=_C_NAVY_DARK)
        sec.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        sec.border    = Border(bottom=Side(style="thin", color=_C_NAVY))
        ws.row_dimensions[row].height = 22
        row += 1

        # Infer payment months from historical timeline
        def _infer_frequency(payment_months: list[int]) -> int | None:
            if len(payment_months) < 2:
                return None
            unique = sorted(set(payment_months))
            if len(unique) < 2:
                return None
            gaps = [unique[i + 1] - unique[i] for i in range(len(unique) - 1)]
            gaps.append(12 - unique[-1] + unique[0])
            counter = Counter(gaps)
            most_common_gap = counter.most_common(1)[0][0]
            if most_common_gap in (1, 2, 3, 4, 6, 12):
                return most_common_gap
            return None

        ticker_months: dict[str, list[int]] = defaultdict(list)
        for r_entry in dividend_timeline:
            month_num = int(r_entry["month"][5:7])
            ticker_months[r_entry["ticker"]].append(month_num)

        today = pd.Timestamp.today()
        future_months: list[tuple[int, int]] = []
        for i in range(1, 13):
            m = today.year * 12 + (today.month - 1) + i
            future_months.append((m // 12, m % 12 + 1))

        # Build calendar data
        calendar_amounts: dict[str, dict[str, float]] = {}
        for ticker in tickers:
            payments = ticker_months.get(ticker, [])
            freq = _infer_frequency(payments)
            if freq is None or len(payments) < 2:
                continue

            typical_months = sorted(set(payments))
            payments_per_year = len(typical_months)
            fund = fund_map.get(ticker, {})
            div_rate = fund.get("Dividend Rate")
            from src.portfolio import get_split_factor
            total_shares = sum(
                lot["shares"] * get_split_factor(ticker, lot.get("purchase_date"))
                for lot in portfolio.get(ticker, [])
            )

            if not (div_rate and div_rate > 0 and total_shares > 0):
                continue

            div_ccy = fund.get("Financial Currency") or get_ticker_currency(ticker)
            fx_rate, _ = get_fx_rate(div_ccy, currency)
            per_payment = round(div_rate / payments_per_year * total_shares * fx_rate, 2)

            amounts: dict[str, float] = {}
            for y, m in future_months:
                key = f"{y}-{m:02d}"
                amounts[key] = per_payment if m in typical_months else 0.0
            calendar_amounts[ticker] = amounts

        if calendar_amounts:
            # Month headers
            month_keys = [f"{y}-{m:02d}" for y, m in future_months]
            cal_headers = ["Ticker"] + [
                cast(pd.Timestamp, pd.Timestamp(f"{y}-{m:02d}-01")).strftime("%b %Y") for y, m in future_months
            ] + ["Annual Total"]
            _write_headers(ws, cal_headers, start_row=row)
            cal_header_row = row
            row += 1

            cal_tickers = sorted(calendar_amounts.keys())
            for ticker in cal_tickers:
                alt = _row_fill(row)
                ws.cell(row, 1, ticker).font = Font(bold=True, size=10)
                ws.cell(row, 1).border = _CELL_BORDER
                if alt:
                    ws.cell(row, 1).fill = alt

                annual = 0.0
                for ci, key in enumerate(month_keys, 2):
                    amt = calendar_amounts[ticker].get(key, 0.0)
                    cell = ws.cell(row, ci)
                    if amt > 0:
                        cell.value = round(amt, 2)
                        cell.number_format = curr_fmt
                        cell.fill = _GREEN_FILL
                    else:
                        cell.value = None
                    cell.font = Font(size=10)
                    cell.border = _CELL_BORDER
                    cell.alignment = Alignment(horizontal="center")
                    annual += amt

                total_cell = ws.cell(row, len(month_keys) + 2, round(annual, 2))
                total_cell.number_format = curr_fmt
                total_cell.font = Font(bold=True, size=10)
                total_cell.border = _CELL_BORDER
                if alt:
                    total_cell.fill = alt
                ws.row_dimensions[row].height = 18
                row += 1

            # Monthly totals row
            totals_row = row
            ws.cell(totals_row, 1, "Monthly Total").font = _TOTAL_FONT
            grand_total = 0.0
            for ci, key in enumerate(month_keys, 2):
                month_sum = sum(
                    calendar_amounts[t].get(key, 0.0) for t in cal_tickers
                )
                c = ws.cell(totals_row, ci, round(month_sum, 2) if month_sum > 0 else None)
                c.number_format = curr_fmt
                c.font = _TOTAL_FONT
                c.fill = _TOTAL_FILL
                c.border = _TOP_BORDER
                c.alignment = Alignment(horizontal="center")
                grand_total += month_sum

            ws.cell(totals_row, 1).fill = _TOTAL_FILL
            ws.cell(totals_row, 1).border = _TOP_BORDER
            gt = ws.cell(totals_row, len(month_keys) + 2, round(grand_total, 2))
            gt.number_format = curr_fmt
            gt.font = _TOTAL_FONT
            gt.fill = _TOTAL_FILL
            gt.border = _TOP_BORDER
            ws.row_dimensions[totals_row].height = 22
            row += 1

            _add_table(ws, "tblDivCalendar",
                       f"A{cal_header_row}:{get_column_letter(len(cal_headers))}{totals_row - 1}")

            # Note
            note = ws.cell(row, 1,
                "Projected from historical payment patterns and current dividend rates. "
                "Actual payments may differ. Tickers with fewer than 2 historical payments are excluded.")
            note.font = _NOTE_FONT
            note.alignment = Alignment(wrap_text=True, indent=1)
            ws.merge_cells(f"A{row}:{get_column_letter(len(cal_headers))}{row}")
            ws.row_dimensions[row].height = 28

    _autofit(ws)


# ── Public API ─────────────────────────────────────────────────────────────────

def build_excel_report(
    positions_df: pd.DataFrame,
    analytics_df: pd.DataFrame,
    fund_rows: list[dict],
    price_histories: dict[str, pd.DataFrame],
    name_map: dict[str, str],
    currency: str,
    summary_kpis: dict,
    bt_result: dict | None = None,
    ticker_mc_results: dict | None = None,
    portfolio_mc: dict | None = None,
    target_prices: dict[str, float | None] | None = None,
    dividend_timeline: list[dict] | None = None,
    portfolio: dict | None = None,
    health_score: dict | None = None,
    health_findings: list[dict] | None = None,
    health_sector_weights: dict[str, float] | None = None,
    health_ticker_sector: dict[str, str] | None = None,
) -> bytes:
    """
    Build a comprehensive multi-sheet interactive Excel report.
    Returns raw bytes suitable for st.download_button(data=...).
    """
    wb = Workbook()
    if wb.active is not None:
        wb.remove(wb.active)

    _bt  = bt_result          or {}
    _tmc = ticker_mc_results  or {}
    _pmc = portfolio_mc       or {}

    builders = [
        ("Net Worth",         lambda: _sheet_net_worth(wb, summary_kpis, currency)),
        ("Summary",           lambda: _sheet_summary(wb, summary_kpis, currency, len(positions_df))),
        ("Positions",         lambda: _sheet_positions(wb, positions_df, name_map, currency, target_prices=target_prices)),
        ("Attribution",       lambda: _sheet_attribution(wb, positions_df, name_map)),
        ("Allocation",        lambda: _sheet_allocation(wb, positions_df, name_map, currency)),
        ("Income",            lambda: _sheet_income(wb, positions_df, fund_rows, name_map, currency,
                                                    dividend_timeline=dividend_timeline, portfolio=portfolio)),
        ("Risk Metrics",      lambda: _sheet_risk(wb, analytics_df, name_map, positions_df, currency)),
        ("Health Score",      lambda: _sheet_health(wb, health_score or {}, health_findings or [],
                                                    health_sector_weights or {}, health_ticker_sector or {})),
        ("Fundamentals",      lambda: _sheet_fundamentals(wb, fund_rows, name_map)),
        ("Correlation",       lambda: _sheet_correlation(wb, price_histories, positions_df)),
        ("Monte Carlo",       lambda: _sheet_monte_carlo(wb, _bt, _tmc, _pmc, name_map, currency)),
        ("Scenario Analysis", lambda: _sheet_scenario(wb, positions_df, name_map, currency)),
        ("Currency Exposure", lambda: _sheet_currency_exposure(wb, positions_df, name_map, currency)),
        ("Price History",     lambda: _sheet_price_history(wb, price_histories)),
        ("Daily Returns",     lambda: _sheet_daily_returns(wb, price_histories)),
        ("Other Assets",      lambda: _sheet_other_assets(wb, currency)),
    ]

    for sheet_name, builder in builders:
        try:
            builder()
        except Exception as exc:
            import traceback
            traceback.print_exc()
            # If the sheet was partially created, reuse it; otherwise create new
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                # Clear partial content
                for row in ws.iter_rows():
                    for cell in row:
                        cell.value = None
            else:
                ws = wb.create_sheet(sheet_name)
            ws["A1"] = f"Error generating sheet: {exc}"
            print(f"[excel_export] Sheet '{sheet_name}' failed: {exc}")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
